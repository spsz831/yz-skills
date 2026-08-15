#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""书签库智能审计：分层采样 + LLM 汇总，输出「书签健康报告」。

用法:
  python ai_audit.py --dir . --out audit.md           # 默认报告
  python ai_audit.py --dir . --samples 3              # 每表抽 3 条代表性行
  python ai_audit.py --dir . --focus 类别名            # 只审指定类别

分层采样控制 LLM 调用量（全库 1800+ 条 → 1~3 次调用）:
  1. 库级统计摘要（纯 Python，0 次调用）→ 拼进单次 prompt
  2. 每表抽 1-2 条代表性行（低频类别优先，覆盖稀疏处）
  3. --focus 专项追问（按需额外调用）

LLM key 从环境变量 AI_API_KEY 读（复用 ai_enrich.call_llm_one）。
"""
import sys, io, os, glob, argparse
from collections import Counter
from datetime import datetime

if getattr(sys.stdout, 'encoding', '').lower() != 'utf-8':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
try:
    from config import COLUMNS, COLUMN_COUNT, TABLE_GLOB, is_index
except ImportError:
    COLUMNS = {'序号': 0, '名称': 1, 'URL': 2, '类别': 3, '网站类型': 4,
               '功能定位': 5, '是否重复': 6, '备注': 7, '添加日期': 8}
    COLUMN_COUNT = 9
    TABLE_GLOB = '*书签汇总.xlsx'
    is_index = lambda x: False

I_NAME = COLUMNS.get('名称', 1)
I_URL = COLUMNS.get('URL', 2)
I_CAT = COLUMNS.get('类别', 3)
I_TYPE = COLUMNS.get('网站类型', 4)
I_DESC = COLUMNS.get('功能定位', 5)
I_DUP = COLUMNS.get('是否重复', 6)
I_NOTE = COLUMNS.get('备注', 7)
I_DATE = COLUMNS.get('添加日期', 8)


def build_audit_summary(dirpath):
    """纯 Python 扫描全库，返回统计 dict（0 次 LLM 调用）。"""
    import openpyxl
    stats = {'tables': 0, 'entries': 0, 'cats': Counter(), 'types': Counter(),
             'dup': 0, 'no_date': 0, 'no_type': 0, 'no_cat': 0, 'no_desc': 0,
             'no_url': 0, 'per_table': {}}
    for f in sorted(glob.glob(os.path.join(dirpath, TABLE_GLOB))):
        name = os.path.basename(f)
        if name.startswith('~$') or '.bak.' in name or is_index(f):
            continue
        wb = openpyxl.load_workbook(f, read_only=True)
        n = 0
        for ws in wb.worksheets:
            for r in ws.iter_rows(min_row=2, values_only=True):
                if r[I_NAME] is None or str(r[I_NAME]).strip() == '':
                    continue
                n += 1
                stats['entries'] += 1
                cat = str(r[I_CAT] or '').strip()
                typ = str(r[I_TYPE] or '').strip()
                desc = str(r[I_DESC] or '').strip()
                if cat:
                    stats['cats'][cat] += 1
                else:
                    stats['no_cat'] += 1
                if typ:
                    stats['types'][typ] += 1
                else:
                    stats['no_type'] += 1
                if not desc:
                    stats['no_desc'] += 1
                if not r[I_URL]:
                    stats['no_url'] += 1
                if str(r[I_DUP] or '') == '是':
                    stats['dup'] += 1
                if not str(r[I_DATE] or '').strip():
                    stats['no_date'] += 1
        stats['tables'] += 1
        stats['per_table'][name.replace('书签汇总.xlsx', '')] = n
        wb.close()
    return stats


def build_sampled_rows(dirpath, per_table=2):
    """每表抽 per_table 条代表性行（低频类别优先，覆盖稀疏处）。"""
    import openpyxl
    from collections import defaultdict
    by_cat = defaultdict(list)
    for f in sorted(glob.glob(os.path.join(dirpath, TABLE_GLOB))):
        name = os.path.basename(f)
        if name.startswith('~$') or '.bak.' in name or is_index(f):
            continue
        wb = openpyxl.load_workbook(f, read_only=True)
        for ws in wb.worksheets:
            for r in ws.iter_rows(min_row=2, values_only=True):
                if r[I_NAME] is None or str(r[I_NAME]).strip() == '':
                    continue
                cat = str(r[I_CAT] or '未分类').strip()
                by_cat[cat].append({
                    't': name.replace('书签汇总.xlsx', ''),
                    'n': str(r[I_NAME]).strip(),
                    'u': str(r[I_URL] or '').strip(),
                    'c': cat, 'y': str(r[I_TYPE] or '').strip(),
                    'd': str(r[I_DESC] or '').strip(),
                })
        wb.close()
    # 低频类别优先：类内条数升序
    cats = sorted(by_cat, key=lambda c: len(by_cat[c]))
    sampled = []
    for c in cats:
        rows = by_cat[c]
        # 每个低频类抽 1 条，够配额为止
        for _ in range(min(per_table, len(rows))):
            sampled.append(rows[0])
    return sampled[: max(per_table * 5, 8)]


def build_prompt(stats, sampled, focus=''):
    """构造审计 prompt（一次调用覆盖库级 + 样本）。"""
    top_cats = ', '.join(f'{k}({v})' for k, v in stats['cats'].most_common(8))
    top_types = ', '.join(f'{k}({v})' for k, v in stats['types'].most_common(6))
    sample_lines = []
    for e in sampled[:12]:
        sample_lines.append(f"- [{e['t']}] {e['n']} | URL: {e['u']} | 类别: {e['c']} | 类型: {e['y']} | 定位: {e['d'][:40]}")
    focus_line = f'\n专项关注: {focus}' if focus else ''
    return f"""你是书签库审计助手。根据以下库级统计和抽样行，输出中文书签健康报告。

库级统计:
- 表数: {stats['tables']}，书签总数: {stats['entries']}
- 重复标记: {stats['dup']} 条；缺类别 {stats['no_cat']}，缺类型 {stats['no_type']}，缺定位 {stats['no_desc']}，缺日期 {stats['no_date']}，缺URL {stats['no_url']}
- 高频类别 TOP8: {top_cats}
- 高频类型 TOP6: {top_types}

抽样行（低频类别优先）:
{sample_lines}
{focus_line}

请输出:
1. 内容过时/冗余信号（类别重叠、类型笼统、疑似失效）
2. 质量缺口（缺字段最严重的方向）
3. 可合并的类别建议
4. 下一步行动清单（优先级排序）

控制在 400 字内。"""


def main(argv):
    ap = argparse.ArgumentParser(description='书签库智能审计（分层采样 + LLM 报告）')
    ap.add_argument('--dir', default='.', help='书签库目录')
    ap.add_argument('--out', default=None, help='报告输出文件（默认 stdout）')
    ap.add_argument('--samples', type=int, default=2, help='每表抽样行数')
    ap.add_argument('--focus', default='', help='专项追问类别/主题')
    ap.add_argument('--selftest', action='store_true')
    args = ap.parse_args(argv)

    if args.selftest:
        return _selftest()

    from ai_enrich import call_llm_one

    stats = build_audit_summary(args.dir)
    if not stats['entries']:
        print(f'⚠️ {args.dir} 下未找到书签')
        return 1
    sampled = build_sampled_rows(args.dir, per_table=args.samples)
    prompt = build_prompt(stats, sampled, args.focus)

    print(f'库级统计: {stats["tables"]} 表 / {stats["entries"]} 条 / {stats["cats"] and sum(stats["cats"].values())} 类别')
    print(f'抽样 {len(sampled)} 条代表性行，请求 LLM 生成审计报告…')

    ok, resp = call_llm_one(prompt)
    if not ok:
        print(f'✗ LLM 调用失败: {resp}')
        print('（已输出库级统计，LLM 报告未能生成）')
        # 兜底：输出纯 Python 统计
        report = f'# 书签库审计报告（LLM 失败，仅统计）\n\n- 表数: {stats["tables"]}\n- 书签总数: {stats["entries"]}\n- 重复标记: {stats["dup"]}\n- 缺类别: {stats["no_cat"]} / 缺类型: {stats["no_type"]} / 缺定位: {stats["no_desc"]}\n'
        if args.out:
            with open(args.out, 'w', encoding='utf-8') as fh:
                fh.write(report)
        return 1

    report = f'# 书签库审计报告（{datetime.now().strftime("%Y-%m-%d")}）\n\n'
    report += f'- 表数: {stats["tables"]} | 书签总数: {stats["entries"]} | 重复标记: {stats["dup"]}\n'
    report += f'- 缺字段: 类别 {stats["no_cat"]} / 类型 {stats["no_type"]} / 定位 {stats["no_desc"]} / 日期 {stats["no_date"]} / URL {stats["no_url"]}\n\n'
    report += '## AI 审计意见\n\n' + resp.strip() + '\n'
    if args.out:
        with open(args.out, 'w', encoding='utf-8') as fh:
            fh.write(report)
        print(f'✅ 报告已写入 {args.out}')
    else:
        print(report)
    return 0


def _selftest():
    import tempfile, shutil, openpyxl
    ok = True
    tmpdir = tempfile.mkdtemp()
    try:
        # 造两表
        for fname, sheet, rows in [
            ('工具书签汇总.xlsx', '工具', [('工具A', 'https://a.com', '工具', '官网·工具', '做X', '2026-01-01')]),
            ('AI书签汇总.xlsx', 'AI工具', [('模型B', 'https://b.com', 'AI工具', 'GitHub·模型', '跑模型', '2026-02-02'),
                                            ('无类别', 'https://c.com', '', '官网·工具', '没类别', '2026-03-03')]),
        ]:
            p = os.path.join(tmpdir, fname)
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet
            for c, h in enumerate(['序号', '名称', 'URL', '类别', '网站类型', '功能定位', '是否重复', '备注', '添加日期'], start=1):
                ws.cell(1, c).value = h
            for i, (n, u, cat, typ, d, dt) in enumerate(rows, start=2):
                ws.cell(i, 1).value = i - 1
                ws.cell(i, 2).value = n; ws.cell(i, 3).value = u
                ws.cell(i, 4).value = cat; ws.cell(i, 5).value = typ
                ws.cell(i, 6).value = d; ws.cell(i, 8).value = dt
            wb.save(p); wb.close()

        stats = build_audit_summary(tmpdir)
        assert stats['tables'] == 2, stats
        assert stats['entries'] == 3, stats
        assert stats['no_cat'] == 1 and stats['no_desc'] == 0, stats
        assert stats['cats']['AI工具'] == 1

        sampled = build_sampled_rows(tmpdir, per_table=1)
        assert len(sampled) >= 2, sampled  # 2表至少各1条

        prompt = build_prompt(stats, sampled)
        assert 'AI工具' in prompt and '模型B' in prompt and '书签总数' in prompt
        assert prompt.count('http') >= 1
        print('selftest: 通过')
        return 0
    except AssertionError as e:
        print(f'selftest: 失败 — {e}')
        return 1
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)


if __name__ == '__main__':
    sys.exit(main(sys.argv[1:]))
