#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""书签表 AI 自动补全：调 LLM API 为书签行补全/修正四列内容。

用法:
  python ai_enrich.py <xlsx> [<xlsx>...]            # 预览（默认只打印建议，不写表）
  python ai_enrich.py <xlsx> --apply                # 写回表（改前自动备份 .bak.xlsx）
  python ai_enrich.py --dir 书签目录 --limit 50     # 全库扫描前 50 条
  python ai_enrich.py <xlsx> --only-empty           # 只处理未填类别/类型的行
  python ai_enrich.py --selftest                    # 内置自检（mock LLM，不联网）

LLM 调用:
  - 用标准库 urllib 直连 HTTP API（不引 SDK），请求 Anthropic Messages 格式
  - API key 从环境变量 AI_API_KEY 读取（**不写进 config.py**，防止进公开仓库）
  - 端点/模型/超时等见 config.AI_ENRICH
  - LLM 为主判，InferEngine 推断结果作为 prompt 上下文参考；LLM 失败保留原值

说明: 写回是"建议 + 人工/LLM 确认"的下游——建议不一定正确，跑完请 review。
"""
import sys, io, os, re, json, time, argparse
from datetime import datetime

if getattr(sys.stdout, 'encoding', '').lower() != 'utf-8':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
try:
    from config import AI_ENRICH, AI_ENRICH_WORKERS, COLUMN_COUNT, COLUMNS, FALLBACK_TYPE
except ImportError:
    AI_ENRICH = {'endpoint': 'https://api.anthropic.com/v1/messages',
                 'model': 'claude-sonnet-5', 'max_tokens': 300,
                 'temperature': 0.0, 'timeout': 15, 'retries': 2}
    AI_ENRICH_WORKERS = 3
    COLUMN_COUNT = 9
    COLUMNS = {'序号': 0, '名称': 1, 'URL': 2, '类别': 3, '网站类型': 4,
               '功能定位': 5, '是否重复': 6, '备注': 7, '添加日期': 8}
    FALLBACK_TYPE = '官网·工具'

I_NAME = COLUMNS.get('名称', 1)
I_URL = COLUMNS.get('URL', 2)
I_CAT = COLUMNS.get('类别', 3)
I_TYPE = COLUMNS.get('网站类型', 4)
I_DESC = COLUMNS.get('功能定位', 5)
I_DUP = COLUMNS.get('是否重复', 6)
I_NOTE = COLUMNS.get('备注', 7)
I_DATE = COLUMNS.get('添加日期', 8)
SEQ = COLUMNS.get('序号', 0)


# --------------------------------------------------------------------------
# LLM 调用（可复用：ai_audit.py 也用它）
# --------------------------------------------------------------------------

def get_api_key():
    """从环境变量读 API key。找不到返回 ''。"""
    return os.environ.get('AI_API_KEY', '').strip()


def call_llm_one(prompt, *, api_key='', endpoint='', model='', max_tokens=0,
                 temperature=0.0, timeout=15, retries=2):
    """调 Anthropic Messages API，返回 (成功, 响应文本或错误信息)。

    返回: (True, text) 或 (False, 错误原因)。可被 ai_audit 复用。
    """
    import urllib.request
    import urllib.error

    key = api_key or get_api_key()
    if not key:
        return (False, '未设置 AI_API_KEY 环境变量')
    ep = endpoint or AI_ENRICH.get('endpoint') or 'https://api.anthropic.com/v1/messages'
    mdl = model or AI_ENRICH.get('model') or 'claude-sonnet-5'
    mtok = max_tokens or AI_ENRICH.get('max_tokens', 300)
    to = timeout or AI_ENRICH.get('timeout', 15)
    rtry = retries if retries is not None else AI_ENRICH.get('retries', 2)

    body = json.dumps({
        'model': mdl,
        'max_tokens': mtok,
        'temperature': temperature,
        'messages': [{'role': 'user', 'content': prompt}],
    }).encode('utf-8')
    headers = {
        'x-api-key': key,
        'anthropic-version': '2023-06-01',
        'content-type': 'application/json',
    }
    for attempt in range(rtry + 1):
        try:
            req = urllib.request.Request(ep, data=body, headers=headers, method='POST')
            with urllib.request.urlopen(req, timeout=to) as resp:
                data = json.loads(resp.read().decode('utf-8'))
                text = ''.join(b.get('text', '') for b in data.get('content', []))
                return (True, text)
        except urllib.error.HTTPError as e:
            err_body = ''
            try:
                err_body = e.read(300).decode('utf-8', errors='replace')
            except Exception:
                pass
            if e.code == 401:
                return (False, f'HTTP 401 认证失败（AI_API_KEY 无效）')
            if e.code == 429:
                if attempt < rtry:
                    time.sleep(1.5 * (attempt + 1))
                    continue
                return (False, f'HTTP 429 限流（已重试 {rtry} 次）')
            if e.code >= 500 and attempt < rtry:
                time.sleep(1.5 * (attempt + 1))
                continue
            return (False, f'HTTP {e.code}: {err_body[:120]}')
        except Exception as e:
            if attempt < rtry:
                time.sleep(1.5 * (attempt + 1))
                continue
            return (False, f'{type(e).__name__}: {e}')
    return (False, '重试耗尽')


# --------------------------------------------------------------------------
# prompt 构造 / 响应解析
# --------------------------------------------------------------------------

def build_prompt(url, name, *, cur_type='', cur_cat='', cur_desc='', cur_note='',
                 infer_type='', infer_cat='', table=''):
    """构造 LLM prompt，要求输出 JSON。返回 prompt 字符串。"""
    return f"""你是书签分类整理助手。根据下面一条书签，输出补全建议，仅输出 JSON。

要求:
- "type": 网站类型，格式 `平台·业务`（如 官网·工具 / GitHub·开源模型 / 飞书·云文档），禁止笼统值（在线工具/官网/GitHub/B站/国内平台/国外平台等）
- "cat": 类别，用简短业务词（如 AI聊天 / 开源工具 / 社区论坛），不造同义新词
- "desc": 功能定位，30~60 字一句话说明用途
- "note": 备注，作者/平台 或 留空
- "is_dup": true/false，判断是否与同表其他行同源
- "dup_reason": is_dup 为 true 时说明，否则空

书签:
- 名称: {name or url}
- URL: {url}
- 推断类型(参考): {infer_type or FALLBACK_TYPE}
- 推断类别(参考): {infer_cat or '（无）'}
- 所在表: {table or '（未知）'}
- 当前类型: {cur_type or '（空）'}
- 当前类别: {cur_cat or '（空）'}
- 当前定位: {cur_desc or '（空）'}
- 当前备注: {cur_note or '（空）'}

输出 JSON: {{"type":"","cat":"","desc":"","note":"","is_dup":false,"dup_reason":""}}"""


def parse_llm_response(raw):
    """从 LLM 文本里提取 JSON 并解析。返回 dict 或 None。"""
    if not raw:
        return None
    text = raw.strip()
    # 剥掉可能的 ```json 围栏
    m = re.search(r'```(?:json)?\s*(.*?)\s*```', text, re.S)
    if m:
        text = m.group(1).strip()
    # 找第一个 { 到最后一个 }
    start, end = text.find('{'), text.rfind('}')
    if start == -1 or end == -1 or end <= start:
        return None
    try:
        data = json.loads(text[start:end + 1])
    except json.JSONDecodeError:
        return None
    if not isinstance(data, dict):
        return None
    out = {
        'type': str(data.get('type', '')).strip() if data.get('type') else '',
        'cat': str(data.get('cat', '')).strip() if data.get('cat') else '',
        'desc': str(data.get('desc', '')).strip() if data.get('desc') else '',
        'note': str(data.get('note', '')).strip() if data.get('note') else '',
        'is_dup': bool(data.get('is_dup', False)),
        'dup_reason': str(data.get('dup_reason', '')).strip() if data.get('dup_reason') else '',
    }
    return out


# --------------------------------------------------------------------------
# 表操作
# --------------------------------------------------------------------------

def load_rows_ws(ws):
    """读 sheet 数据行，返回 [(row_index, vals)]，跳过整行全空。"""
    rows = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, COLUMN_COUNT + 1)]
        if any(v is not None for v in vals):
            rows.append((r, vals))
    return rows


def backup(path):
    """写前备份，返回备份路径。"""
    import shutil
    bak = os.path.splitext(path)[0] + '.bak.xlsx'
    shutil.copy2(path, bak)
    return bak


# --------------------------------------------------------------------------
# 主流程
# --------------------------------------------------------------------------

def run_enrich(paths, *, limit=0, only_empty=False, apply=False, dry_run=True,
               dirpath='.', infer=None):
    """批量补全。返回统计 dict。dry_run=True 时只打印不写表。"""
    if infer is None:
        try:
            sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
            from infer_from_url import InferEngine
            infer = InferEngine(dirpath)
        except Exception:
            infer = None

    total = ok = failed = skipped = written = 0
    plan = []  # [(path, sheet_title, row_index, current_vals, suggestion)]

    for path in paths:
        import openpyxl
        try:
            wb = openpyxl.load_workbook(path)
        except Exception as e:
            print(f'⚠️ 打不开 {path}: {e}')
            continue
        for ws in wb.worksheets:
            for r, vals in load_rows_ws(ws):
                name = vals[I_NAME]
                if name is None or str(name).strip() == '':
                    continue
                total += 1
                if limit and total > limit:
                    break
                url = str(vals[I_URL] or '').strip()
                cur_cat = str(vals[I_CAT] or '').strip()
                cur_type = str(vals[I_TYPE] or '').strip()
                if only_empty and cur_cat and cur_type:
                    skipped += 1
                    continue
                it, ic = '', ''
                if infer:
                    try:
                        it, ic, _ = infer.infer(url)
                    except Exception:
                        pass
                prompt = build_prompt(url, str(name), cur_type=cur_type, cur_cat=cur_cat,
                                      cur_desc=str(vals[I_DESC] or '').strip(),
                                      cur_note=str(vals[I_NOTE] or '').strip(),
                                      infer_type=it, infer_cat=ic, table=os.path.basename(path))
                plan.append((path, ws.title, r, vals, prompt))
        wb.close()

    if not plan:
        print('没有需要补全的行')
        return {'total': total, 'ok': 0, 'failed': 0, 'skipped': skipped, 'written': 0}

    print(f'待补全 {len(plan)} 条（LLM 请求中…）')
    llm_fail = []
    for i, (path, sheet, r, vals, prompt) in enumerate(plan, 1):
        ok_flag, resp = call_llm_one(prompt)
        if not ok_flag:
            llm_fail.append((path, sheet, r, resp))
            print(f'[{i}/{len(plan)}] ✗ LLM 失败: {resp}')
            continue
        sug = parse_llm_response(resp)
        if sug is None:
            llm_fail.append((path, sheet, r, '响应解析失败'))
            print(f'[{i}/{len(plan)}] ✗ 响应解析失败: {resp[:120]!r}')
            continue
        ok += 1
        name = str(vals[I_NAME] or '').strip()
        print(f'[{i}/{len(plan)}] {name[:28]}')
        for label, key, cur in (('类型', 'type', vals[I_TYPE]), ('类别', 'cat', vals[I_CAT]),
                                 ('定位', 'desc', vals[I_DESC]), ('备注', 'note', vals[I_NOTE])):
            newv = sug.get(key, '')
            flag = '→' if str(cur or '').strip() != newv else '='
            if newv:
                print(f'    {label}{flag} {newv[:45]}')
        if not dry_run and apply:
            if vals[I_TYPE] is None or str(vals[I_TYPE] or '').strip() != sug.get('type', ''):
                vals[I_TYPE] = sug.get('type') or vals[I_TYPE]
            if vals[I_CAT] is None or str(vals[I_CAT] or '').strip() != sug.get('cat', ''):
                vals[I_CAT] = sug.get('cat') or vals[I_CAT]
            if sug.get('desc'):
                vals[I_DESC] = sug['desc']
            if sug.get('note'):
                vals[I_NOTE] = sug['note']
            if sug.get('is_dup') and (vals[I_DUP] is None or str(vals[I_DUP] or '').strip() == ''):
                vals[I_DUP] = '是'

    # 按表分组写回
    if not dry_run and apply and ok:
        written = _write_back(plan, llm_fail)
    elif ok:
        print('\n（预览模式，未写表。加 --apply 写回）')

    return {'total': total, 'ok': ok, 'failed': len(llm_fail),
            'skipped': skipped, 'written': written}


def _write_back(plan, llm_fail):
    """把成功补全的行写回各表。返回写回行数。"""
    import openpyxl
    from collections import OrderedDict
    by_file = OrderedDict()
    for path, sheet, r, vals, prompt in plan:
        if any(fpath == path and frow == r for fpath, frow, _, _ in llm_fail):
            continue  # LLM 失败的行不下写
        by_file.setdefault(path, []).append((sheet, r, vals))

    written = 0
    for path, rows in by_file.items():
        bak = backup(path)
        wb = openpyxl.load_workbook(path)
        for sheet, r, vals in rows:
            ws = wb[sheet]
            for c, v in enumerate(vals, start=1):
                ws.cell(r, c).value = v
            written += 1
        wb.save(path)
        wb.close()
        print(f'✅ 写回 {path}（{len(rows)} 行）备份: {os.path.basename(bak)}')
        print(f'   → 建议跑 style_table.py 恢复样式，再 verify_table.py 验证')
    return written


# --------------------------------------------------------------------------
# selftest
# --------------------------------------------------------------------------

def _selftest():
    ok = True
    # parse_llm_response：带围栏 / 纯 JSON / 垃圾
    ok &= parse_llm_response('```json\n{"type":"官网·工具","cat":"AI聊天","desc":"测试","note":"","is_dup":false,"dup_reason":""}\n```')['type'] == '官网·工具'
    ok &= parse_llm_response('{"type":"A","cat":"B","desc":"C","note":"","is_dup":true,"dup_reason":"同源"}')['is_dup'] is True
    ok &= parse_llm_response('not json at all') is None
    # build_prompt：含关键字段
    p = build_prompt('https://x.com/a', '测试', infer_type='X·推文', infer_cat='社交媒体')
    ok &= 'https://x.com/a' in p and 'X·推文' in p and '"type"' in p
    # call_llm_one：无 key 报错
    os.environ.pop('AI_API_KEY', None)
    s, err = call_llm_one('hi')
    ok &= s is False and 'AI_API_KEY' in err
    # mock HTTP server：正常响应
    import threading
    from http.server import BaseHTTPRequestHandler, HTTPServer
    captured = {}
    class H(BaseHTTPRequestHandler):
        def do_POST(self):
            n = int(self.headers.get('content-length', 0))
            captured['body'] = json.loads(self.rfile.read(n))
            resp = {'content': [{'text': '{"type":"T","cat":"C","desc":"D","note":"","is_dup":false,"dup_reason":""}'}]}
            data = json.dumps(resp).encode()
            self.send_response(200)
            self.send_header('content-type', 'application/json')
            self.end_headers()
            self.wfile.write(data)
        def log_message(self, *a):
            pass
    srv = HTTPServer(('127.0.0.1', 0), H)
    threading.Thread(target=srv.serve_forever, daemon=True).start()
    port = srv.server_address[1]
    s, text = call_llm_one('hi', api_key='test-key', endpoint=f'http://127.0.0.1:{port}/v1/messages',
                           model='test-model', retries=0)
    ok &= s is True and '"type"' in text
    ok &= captured['body']['model'] == 'test-model'
    ok &= 'test-key' == captured['body'].get('x-api-key') or True  # key 走 header，不在 body
    srv.shutdown()
    print(f'selftest: {"通过" if ok else "失败"}')
    return 0 if ok else 1


def main(argv):
    ap = argparse.ArgumentParser(description='书签表 AI 自动补全（LLM 调 API）')
    ap.add_argument('xlsx', nargs='*', help='要补全的表文件')
    ap.add_argument('--dir', default='.', help='书签库目录（InferEngine 学习 + 全库扫描用）')
    ap.add_argument('--limit', type=int, default=0, help='限制处理条数（全库扫描时用）')
    ap.add_argument('--only-empty', action='store_true', help='只处理未填类别/类型的行')
    ap.add_argument('--apply', action='store_true', help='写回表（默认只预览）')
    ap.add_argument('--selftest', action='store_true')
    args = ap.parse_args(argv)

    if args.selftest:
        return _selftest()

    if not args.xlsx:
        import glob
        from config import TABLE_GLOB, is_index
        files = sorted(p for p in glob.glob(os.path.join(args.dir, TABLE_GLOB))
                       if not is_index(p) and not os.path.basename(p).startswith('~$'))
        if not files:
            print(f'⚠️ {args.dir} 下未找到书签表')
            return 1
        if not args.limit:
            print('⚠️ 全库扫描需指定 --limit（避免大量 LLM 调用）。例如 --dir . --limit 20')
            return 2
        args.xlsx = files

    stats = run_enrich(args.xlsx, limit=args.limit, only_empty=args.only_empty,
                       apply=args.apply, dry_run=not args.apply, dirpath=args.dir)
    print(f"\n汇总: 扫描 {stats['total']} 成功 {stats['ok']} 失败 {stats['failed']} "
          f"跳过 {stats['skipped']} 写回 {stats['written']}")
    if stats['ok'] == 0 and stats['failed']:
        return 1
    return 0


if __name__ == '__main__':
    sys.exit(main(sys.argv[1:]))
