#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""书签检索门户：全库 → 单文件可搜索 HTML（零依赖，双击即用）。

用法:
  python build_portal.py --dir . --out portal.html

输出: 单文件 HTML，内嵌 CSS + 原生 JS：
  - 搜索框：按名称/URL/类别/定位/备注 实时过滤
  - 类别标签栏：点击切换
  - 按类别分区块，每区按表分组展示；顶部统计（共 N 条 | M 类别 | K 表）
  - 本地 file:// 直接打开，无外部依赖

说明: 只读，不写任何表。数据嵌入 HTML，重新生成即可刷新。
"""
import sys, io, os, glob, json, argparse
from html import escape

if getattr(sys.stdout, 'encoding', '').lower() != 'utf-8':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
try:
    from config import COLUMNS, COLUMN_COUNT, TABLE_GLOB, is_index
except ImportError:
    COLUMNS = {'序号': 0, '名称': 1, 'URL': 2, '类别': 3, '网站类型': 4,
               '功能定位': 5, '是否重复': 6, '备注': 7, '添加日期': 8}
    COLUMN_COUNT = 9
    TABLE_GLOB = '*书签汇总.xlsx'
    is_index = lambda x: False

I_NAME = COLUMNS.get('名称', 1)
I_URL = COLUMNS.get('URL', 2)
I_CAT = COLUMNS.get('类别', 3)
I_TYPE = COLUMNS.get('网站类型', 4)
I_DESC = COLUMNS.get('功能定位', 5)
I_DUP = COLUMNS.get('是否重复', 6)
I_NOTE = COLUMNS.get('备注', 7)
I_DATE = COLUMNS.get('添加日期', 8)


def collect_all_entries(dirpath):
    """读全库所有书签行（含多 sheet），返回 [dict]。跳过空行/索引表/备份。"""
    import openpyxl
    entries = []
    files = sorted(p for p in glob.glob(os.path.join(dirpath, TABLE_GLOB))
                   if not os.path.basename(p).startswith('~$') and '.bak.' not in p
                   and not is_index(p))
    for path in files:
        wb = openpyxl.load_workbook(path, read_only=True)
        table = os.path.basename(path).replace('书签汇总.xlsx', '')
        for ws in wb.worksheets:
            for r in ws.iter_rows(min_row=2, values_only=True):
                name = r[I_NAME]
                if name is None or str(name).strip() == '':
                    continue
                url = str(r[I_URL] or '').strip()
                if not url:
                    continue
                entries.append({
                    't': table, 's': ws.title,
                    'n': str(name).strip(),
                    'u': url,
                    'c': str(r[I_CAT] or '').strip(),
                    'y': str(r[I_TYPE] or '').strip(),
                    'd': str(r[I_DESC] or '').strip(),
                    'k': str(r[I_NOTE] or '').strip(),
                    'a': str(r[I_DATE] or '').strip(),
                })
        wb.close()
    return entries


def group_by_category(entries):
    """按类别分组，返回 {类别: [entry,...]}（保持首现顺序）。空类别归 '未分类'。"""
    groups = {}
    for e in entries:
        cat = e['c'] or '未分类'
        groups.setdefault(cat, []).append(e)
    return groups


def render_html(entries, dirpath, now):
    """渲染单文件可搜索 HTML。数据序列化进 JSON 供 JS 过滤。"""
    groups = group_by_category(entries)
    cats = sorted(groups.keys())
    # 防 </script> 注入：把 / 转义，JSON 解析回来是原值
    data = json.dumps(entries, ensure_ascii=False).replace('</', '<\\/')

    n_tables = len({e['t'] for e in entries})
    n_cats = len(cats)
    n_dup = sum(1 for e in entries if e.get('_d', False))

    # 内嵌样式 + 原生 JS
    css = """
    * { box-sizing: border-box; margin: 0; padding: 0; }
    body { font-family: "Microsoft YaHei", sans-serif; background: #f5f6f8; color: #222; padding: 20px; }
    header { max-width: 1100px; margin: 0 auto 16px; }
    h1 { font-size: 22px; color: #1a3e6e; }
    .stats { color: #666; font-size: 13px; margin-top: 4px; }
    #q { width: 100%; max-width: 1100px; padding: 10px 14px; font-size: 15px;
         border: 1px solid #ccc; border-radius: 6px; display: block; margin: 0 auto 12px; }
    .cats { max-width: 1100px; margin: 0 auto 12px; display: flex; flex-wrap: wrap; gap: 6px; }
    .cat { padding: 4px 12px; border-radius: 999px; border: 1px solid #ccc; background: #fff;
           cursor: pointer; font-size: 13px; }
    .cat.active { background: #1a3e6e; color: #fff; border-color: #1a3e6e; }
    main { max-width: 1100px; margin: 0 auto; }
    .group { margin-bottom: 24px; }
    .group h2 { font-size: 16px; color: #1a3e6e; border-left: 4px solid #1a3e6e;
                padding-left: 8px; margin-bottom: 8px; }
    .tbl { font-size: 13px; color: #888; margin-bottom: 6px; }
    .items { display: grid; grid-template-columns: repeat(auto-fill, minmax(320px, 1fr)); gap: 8px; }
    .card { background: #fff; border: 1px solid #e3e6ea; border-radius: 6px; padding: 10px 12px; }
    .card a { color: #2563eb; text-decoration: none; font-weight: 600; font-size: 14px; }
    .card a:hover { text-decoration: underline; }
    .card .meta { color: #777; font-size: 12px; margin-top: 4px; }
    .card .desc { color: #444; font-size: 13px; margin-top: 6px; line-height: 1.5; }
    .card .tag { display: inline-block; background: #eef4fb; color: #1a3e6e; border-radius: 4px;
                 padding: 1px 8px; font-size: 12px; margin: 4px 6px 0 0; }
    .card .dup { background: #fdeaea; color: #c0392b; }
    .empty { color: #999; padding: 30px; text-align: center; }
    @media (max-width: 700px) { .items { grid-template-columns: 1fr; } }
    """

    js = """
    var DATA = __DATA__;
    var catSel = null, q = '';

    function render() {
        var kw = q.toLowerCase();
        var hit = DATA.filter(function(e) {
            if (catSel && e.c !== catSel) return false;
            if (!kw) return true;
            return (e.n + ' ' + e.u + ' ' + e.c + ' ' + e.y + ' ' + e.d + ' ' + e.k)
                .toLowerCase().indexOf(kw) !== -1;
        });
        var groups = {};
        hit.forEach(function(e) { (groups[e.c || '未分类'] = groups[e.c || '未分类'] || []).push(e); });
        var cats = Object.keys(groups).sort();
        var html = '';
        if (!cats.length) { html = '<div class="empty">无匹配书签</div>'; }
        cats.forEach(function(c) {
            html += '<div class="group"><h2>' + c + '（' + groups[c].length + '）</h2>';
            var byTable = {};
            groups[c].forEach(function(e) {
                var k = e.t + (e.s !== e.t ? '/' + e.s : '');
                (byTable[k] = byTable[k] || []).push(e);
            });
            Object.keys(byTable).sort().forEach(function(k) {
                html += '<div class="tbl">' + k + '</div><div class="items">';
                byTable[k].forEach(function(e) {
                    var dup = e.c && (e.d === '是') ? ' <span class="tag dup">重复</span>' : '';
                    html += '<div class="card"><a href="' + e.u + '" target="_blank">'
                        + e.n + '</a>' + dup + '<div class="meta">' + e.u + ' · ' + e.a + '</div>';
                    if (e.y) html += '<span class="tag">' + e.y + '</span>';
                    if (e.d) html += '<span class="tag">' + e.d + '</span>';
                    if (e.desc || e.k) html += '<div class="desc">' + e.desc
                        + (e.k ? '<br><span style="color:#888">' + e.k + '</span>' : '') + '</div>';
                    html += '</div>';
                });
                html += '</div>';
            });
            html += '</div>';
        });
        document.getElementById('list').innerHTML = html;
        document.getElementById('cnt').textContent = hit.length + ' / ' + DATA.length;
    }

    window.onload = function() {
        document.getElementById('q').addEventListener('input', function() {
            q = this.value; render();
        });
        var bar = document.getElementById('cats');
        DATA.reduce(function(a, e) { var c = e.c || '未分类'; if (a.indexOf(c) < 0) a.push(c); return a; }, [])
            .sort().forEach(function(c) {
                var el = document.createElement('div');
                el.className = 'cat'; el.textContent = c;
                el.onclick = function() {
                    document.querySelectorAll('.cat').forEach(function(x) { x.classList.remove('active'); });
                    el.classList.add('active');
                    catSel = (catSel === c) ? null : c;
                    if (catSel === null) el.classList.remove('active');
                    render();
                };
                bar.appendChild(el);
            });
        render();
    };
    """

    body = []
    body.append('<!DOCTYPE html>')
    body.append('<html lang="zh"><head><meta charset="utf-8">')
    body.append('<meta name="viewport" content="width=device-width, initial-scale=1">')
    body.append(f'<title>书签门户 · {n_tables} 表 {n_cats} 类 {len(entries)} 条</title>')
    body.append(f'<style>{css}</style></head><body>')
    body.append('<header><h1>📚 书签门户</h1>')
    body.append(f'<div class="stats">共 <b id="cnt">{len(entries)}</b> 条 | {n_cats} 类别 | {n_tables} 表'
                f' | 更新于 {now} | 数据源 {os.path.abspath(dirpath)}</div></header>')
    body.append('<input id="q" type="search" placeholder="搜索名称 / URL / 类别 / 网站类型 / 定位 / 备注…">')
    body.append('<div class="cats" id="cats"></div>')
    body.append('<main id="list"></main>')
    body.append(f'<script>{js.replace("__DATA__", data)}</script>')
    body.append('</body></html>')
    return '\n'.join(body)


def main(argv):
    ap = argparse.ArgumentParser(description='书签检索门户（全库 → 单文件可搜索 HTML）')
    ap.add_argument('--dir', default='.', help='书签库目录')
    ap.add_argument('--out', default='portal.html', help='输出 HTML 路径')
    ap.add_argument('--selftest', action='store_true')
    args = ap.parse_args(argv)

    if args.selftest:
        return _selftest()

    entries = collect_all_entries(args.dir)
    if not entries:
        print(f'⚠️ {args.dir} 下未找到书签')
        return 1
    html = render_html(entries, args.dir, __import__('datetime').datetime.now().strftime('%Y-%m-%d'))
    with open(args.out, 'w', encoding='utf-8') as fh:
        fh.write(html)
    tables = len({e['t'] for e in entries})
    cats = len(group_by_category(entries))
    print(f'✅ 已生成 {args.out}: {len(entries)} 条 | {cats} 类别 | {tables} 表')
    return 0


def _selftest():
    import tempfile, shutil, openpyxl
    ok = True
    tmpdir = tempfile.mkdtemp()
    try:
        p1 = os.path.join(tmpdir, 'AI书签汇总.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'AI工具'  # 默认 sheet 改名，避免删空 sheet
        for c, h in enumerate(['序号', '名称', 'URL', '类别', '网站类型', '功能定位', '是否重复', '备注', '添加日期'], start=1):
            ws.cell(1, c).value = h
        ws.cell(2, 1).value = 1; ws.cell(2, 2).value = '工具A'; ws.cell(2, 3).value = 'https://a.com'
        ws.cell(2, 4).value = 'AI工具'; ws.cell(2, 5).value = '官网·工具'; ws.cell(2, 6).value = '做某事的工具'
        ws.cell(2, 7).value = '是'; ws.cell(2, 8).value = '作者'; ws.cell(2, 9).value = '2026-01-01'
        ws2 = wb.create_sheet('大模型')
        for c, h in enumerate(['序号', '名称', 'URL', '类别', '网站类型', '功能定位', '是否重复', '备注', '添加日期'], start=1):
            ws2.cell(1, c).value = h
        ws2.cell(2, 1).value = 1; ws2.cell(2, 2).value = '模型B'; ws2.cell(2, 3).value = 'https://b.com'
        ws2.cell(2, 4).value = 'AI工具'; ws2.cell(2, 5).value = 'GitHub·模型'
        wb.save(p1); wb.close()

        entries = collect_all_entries(tmpdir)
        assert len(entries) == 2, entries
        assert entries[0]['t'] == 'AI' and entries[0]['s'] == 'AI工具'
        assert entries[0]['c'] == 'AI工具' and entries[0]['k'] == '作者'
        groups = group_by_category(entries)
        assert 'AI工具' in groups and len(groups['AI工具']) == 2

        html = render_html(entries, tmpdir, '2026-08-16')
        assert '书签门户' in html and '工具A' in html and 'https://a.com' in html
        # 数据区无裸 </script>（注入防护：</ 已转义为 <\/）
        data_seg = html.split('var DATA = ', 1)[1].split(';\n', 1)[0]
        assert '</script>' not in data_seg, '数据区存在未转义 </script>'
        assert 'var DATA = ' in html
        print('selftest: 通过')
        return 0
    except AssertionError as e:
        print(f'selftest: 失败 — {e}')
        return 1
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)


if __name__ == '__main__':
    sys.exit(main(sys.argv[1:]))
