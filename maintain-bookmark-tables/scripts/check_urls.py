#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""书签 URL 健康检查：并发 HEAD/GET 检测死链 / 服务器错误 / 连接失败。只读。

用法:
  python check_urls.py <xlsx> [<xlsx>...]            # 检查指定表
  python check_urls.py --dir 书签目录 --limit 100      # 全库抽样前 100 条
  python check_urls.py <xlsx> --skip example.com,x    # 跳过域名防误报

判定:
  2xx/3xx 健康；404/410 死链；其他 4xx 客户端错误；5xx 服务器错误；ERR 连接失败。
  HEAD 被拒(405/403/501) 时回落 GET。localhost/私有网段/非 http(s) 自动跳过。

注意: 会真实发起网络请求；部分站点限流/反爬，--skip 可排除。默认并发 8。
"""
import sys, os, re, socket, ipaddress, argparse
from concurrent.futures import ThreadPoolExecutor, as_completed
from urllib.request import Request, urlopen
from urllib.error import HTTPError, URLError
from urllib.parse import urlparse
from collections import defaultdict

if getattr(sys.stdout, 'encoding', '').lower() != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')
try:
    from config import HTTP_TIMEOUT, HTTP_WORKERS, HTTP_UA, TABLE_GLOB, COLUMN_COUNT, COLUMNS
except ImportError:
    HTTP_TIMEOUT = 8
    HTTP_WORKERS = 8
    HTTP_UA = ('Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 '
               '(KHTML, like Gecko) Chrome/126.0 Safari/537.36')
    TABLE_GLOB = '*书签汇总.xlsx'
    COLUMN_COUNT = 9
    COLUMNS = {'序号': 0, '名称': 1, 'URL': 2, '类别': 3, '网站类型': 4,
               '功能定位': 5, '是否重复': 6, '备注': 7, '添加日期': 8}

TIMEOUT = HTTP_TIMEOUT
WORKERS = HTTP_WORKERS
UA = HTTP_UA

I_NAME = COLUMNS.get('名称', 1)
I_URL = COLUMNS.get('URL', 2)


def skip_host(host, skip_set):
    """跳过本地/私有地址或用户指定的域名。"""
    if not host or host in skip_set:
        return True
    if host in ('localhost', '127.0.0.1', '::1', '0.0.0.0'):
        return True
    if host.endswith('.local') or host.endswith('.localhost'):
        return True
    try:
        if ipaddress.ip_address(host).is_private:
            return True
    except ValueError:
        pass
    return False


def _ascii_url(url):
    """非 ASCII 主机名（中文域名）→ IDNA punycode，urllib 才能发起请求。"""
    p = urlparse(url)
    if not p.hostname or p.hostname.isascii():
        return url
    ascii_host = p.hostname.encode('idna').decode('ascii')
    return p._replace(netloc=p.netloc.replace(p.hostname, ascii_host, 1)).geturl()


def check_one(url, skip_set):
    """返回 (url, 状态) ；状态为 int 状态码 或 'ERR:原因'。"""
    parsed = urlparse(url)
    if parsed.scheme not in ('http', 'https'):
        return (url, 'SKIP')
    host = (parsed.hostname or '').lower()
    if skip_host(host, skip_set):
        return (url, 'SKIP')
    url = _ascii_url(url)

    def _req(method, headers=None):
        return Request(url, headers=headers or {'User-Agent': UA}, method=method)

    try:
        with urlopen(_req('HEAD'), timeout=TIMEOUT) as resp:
            return (url, resp.status)
    except HTTPError as e:
        if e.code in (405, 403, 501):
            try:
                with urlopen(_req('GET', {'User-Agent': UA, 'Range': 'bytes=0-0'}),
                             timeout=TIMEOUT) as resp:
                    return (url, resp.status)
            except HTTPError as e2:
                return (url, e2.code)
            except (URLError, OSError) as e2:
                return (url, f'ERR:{type(e2).__name__}')
        return (url, e.code)
    except (URLError, OSError, socket.timeout) as e:
        return (url, f'ERR:{type(e).__name__}')


def load_entries(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    out = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, I_NAME + 1).value
        if name is None:
            continue
        url = str(ws.cell(r, I_URL + 1).value or '').strip()
        out.append((r, name, url))
    wb.close()
    return out


def main(argv):
    ap = argparse.ArgumentParser(description='书签 URL 健康检查')
    ap.add_argument('xlsx', nargs='*', help='要检查的表文件')
    ap.add_argument('--dir', default='.', help='书签目录（配 --limit 用）')
    ap.add_argument('--limit', type=int, default=0, help='限制检查条数（--dir 全库时必填）')
    ap.add_argument('--skip', default='', help='逗号分隔跳过的域名')
    ap.add_argument('--workers', type=int, default=WORKERS)
    ap.add_argument('--selftest', action='store_true')
    args = ap.parse_args(argv)

    if args.selftest:
        ok = True
        ok &= skip_host('localhost', set())
        ok &= skip_host('192.168.1.10', set())
        ok &= skip_host('10.0.0.5', set())
        ok &= not skip_host('example.com', set())
        ok &= skip_host('example.com', {'example.com'})
        ok &= skip_host('ftp.example.com', set()) is False and True  # 协议判定在 check_one
        ok &= _ascii_url('https://例子.测试/p') == 'https://xn--fsqu00a.xn--0zwm56d/p'
        ok &= _ascii_url('https://example.com/a') == 'https://example.com/a'
        print(f'selftest: {"通过" if ok else "失败"}')
        return 0 if ok else 1

    skip_set = {d.strip().lower() for d in args.skip.split(',') if d.strip()}
    files = list(args.xlsx)
    if not files:
        import glob
        allf = sorted(p for p in glob.glob(os.path.join(args.dir, TABLE_GLOB))
                      if not os.path.basename(p).startswith('~$'))
        if not args.limit:
            print('⚠️ 全库检查需指定 --limit（避免一次 1700+ 请求）。例如 --dir . --limit 100')
            return 2
        files = allf

    entries = []  # (表名, 行号, 名称, url)
    for f in files:
        if not os.path.exists(f):
            print(f'⚠️ 文件不存在: {f}')
            continue
        for r, name, url in load_entries(f):
            entries.append((os.path.basename(f), r, name, url))
    if not entries:
        print('没有可检查的书签')
        return 1
    if args.limit:
        entries = entries[:args.limit]

    print(f'检查 {len(entries)} 条（workers={args.workers}，timeout={TIMEOUT}s）…')
    results = {}
    with ThreadPoolExecutor(max_workers=args.workers) as ex:
        futs = {ex.submit(check_one, url, skip_set): (t, r, n, url)
                for t, r, n, url in entries}
        for fut in as_completed(futs):
            t, r, n, url = futs[fut]
            results[(t, r)] = (fut.result()[1], n, url)

    dead, client, server, err, skip = [], [], [], [], []
    for (t, r), (st, n, url) in sorted(results.items()):
        if isinstance(st, int):
            if st < 400:
                continue
            if st in (404, 410):
                dead.append((t, r, n, url, st))
            elif st < 500:
                client.append((t, r, n, url, st))
            else:
                server.append((t, r, n, url, st))
        elif st == 'SKIP':
            skip.append((t, r, n))
        else:
            err.append((t, r, n, url, st))

    groups = [('死链 404/410', dead), ('客户端错误 4xx', client),
              ('服务器错误 5xx', server), ('连接失败 ERR', err)]
    print()
    for title, items in groups:
        print(f'## {title}（{len(items)}）')
        by_table = defaultdict(list)
        for t, r, n, url, st in items:
            by_table[t].append((r, n, url, st))
        for t, rows in sorted(by_table.items()):
            print(f'- {t}')
            for r, n, url, st in rows:
                print(f'    序号{r} | {str(n)[:30]} | {url[:60]} | {st}')
        if not items:
            print('（无）')
        print()

    total = len(entries)
    healthy = total - len(dead) - len(client) - len(server) - len(err) - len(skip)
    print(f'汇总: 健康 {healthy} / 死链 {len(dead)} / 4xx {len(client)} / '
          f'5xx {len(server)} / 连接失败 {len(err)} / 跳过 {len(skip)} / 共 {total}')
    return 1 if (dead or err) else 0


if __name__ == '__main__':
    sys.exit(main(sys.argv[1:]))
