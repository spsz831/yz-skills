#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""书签条目日常增删改统一入口。

按 URL 定位行，一个命令完成新增 / 修改 / 删除，自动重编号、自动备份、写后提示重跑样式。

用法:
  # 新增：全库自动查重（同 URL 已收录则拒绝，除非 --force）
  python entry.py add "名称" "https://url" --table XX书签汇总.xlsx \
      [--cat 类别] [--desc 功能定位] [--type 网站类型] [--sheet 名] [--infer] [--force]
  # 修改：按 URL 定位，只改指定列
  python entry.py update --table XX书签汇总.xlsx --url <url> --set 类别=新类别,备注=xxx
  # 删除：先 dry-run 预览，确认后执行
  python entry.py delete --table XX书签汇总.xlsx --url <url> [--dry-run] [--yes]

共同行为:
  - 写前自动备份 <表名>.bak.xlsx（--no-backup 关闭）
  - 序号自动重排（依赖序号做引用的表会被破坏，跑完记得 verify_table）
  - 写后提示重跑 style_table.py 恢复美化
"""
import sys, io, os, shutil, glob, argparse
from datetime import datetime

import openpyxl

if getattr(sys.stdout, 'encoding', '').lower() != 'utf-8':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
try:
    from config import COLUMN_COUNT, COLUMNS, TABLE_GLOB, DUP_MARK, DUP_PREFIX
except ImportError:
    COLUMN_COUNT = 9
    COLUMNS = {'序号': 0, '名称': 1, 'URL': 2, '类别': 3, '网站类型': 4,
               '功能定位': 5, '是否重复': 6, '备注': 7, '添加日期': 8}
    TABLE_GLOB = '*书签汇总.xlsx'
    DUP_MARK = '是'
    DUP_PREFIX = '重复：'

I_NAME = COLUMNS.get('名称', 1)
I_URL = COLUMNS.get('URL', 2)
I_CAT = COLUMNS.get('类别', 3)
I_TYPE = COLUMNS.get('网站类型', 4)
I_DESC = COLUMNS.get('功能定位', 5)
I_DUP = COLUMNS.get('是否重复', 6)
I_NOTE = COLUMNS.get('备注', 7)
I_DATE = COLUMNS.get('添加日期', 8)
SEQ = COLUMNS.get('序号', 0)


def norm_url(u):
    return str(u).strip().rstrip('/') if u else ''


def load_rows(ws):
    rows = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, COLUMN_COUNT + 1)]
        if any(v is not None for v in vals):
            rows.append(vals)
    return rows


def write_rows(ws, rows):
    """清空第2行起的旧数据区，重写数据。表头保留。"""
    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row)
    for i, vals in enumerate(rows, start=2):
        for c, v in enumerate(vals, start=1):
            ws.cell(i, c).value = v


def renumber(rows):
    for i, vals in enumerate(rows, start=1):
        vals[SEQ] = i


def find_row_by_url(rows, url):
    target = norm_url(url)
    for i, vals in enumerate(rows):
        if norm_url(vals[I_URL]) == target:
            return i
    return None


def insert_pos(rows, cat, before, after):
    """返回应插入的 list 下标。优先级: --before > --after > 同类别末尾 > 表尾。"""
    if before:
        idx = find_name_row(rows, before)
        if idx is not None:
            return idx
    if after:
        idx = find_name_row(rows, after)
        if idx is not None:
            return idx + 1
    last_same = -1
    for i, vals in enumerate(rows):
        if vals[I_CAT] and cat and str(vals[I_CAT]).strip() == str(cat).strip():
            last_same = i
    return last_same + 1


def find_name_row(rows, name):
    for i, vals in enumerate(rows):
        if vals[I_NAME] and str(vals[I_NAME]).strip() == str(name).strip():
            return i
    return None


def backup(path, enabled):
    """写前备份。返回备份路径或 None。"""
    if not enabled:
        return None
    bak = os.path.splitext(path)[0] + '.bak.xlsx'
    shutil.copy2(path, bak)
    return bak


def list_tables(dirpath):
    """全库书签表文件列表（跳过索引表、跳过备份）。"""
    out = []
    for p in glob.glob(os.path.join(dirpath, TABLE_GLOB)):
        name = os.path.basename(p)
        if name.startswith('~$') or '.bak.' in name:
            continue
        try:
            from config import is_index
        except ImportError:
            is_index = lambda x: False
        if is_index(p):
            continue
        out.append(p)
    return sorted(out)


def find_dup_elsewhere(dirpath, url, exclude):
    """全库扫描同 URL。返回 [(表路径, sheet名, 行号, 名称)]，排除 exclude 表本身。"""
    target = norm_url(url)
    hits = []
    for p in list_tables(dirpath):
        if os.path.abspath(p) == os.path.abspath(exclude):
            continue
        try:
            wb = openpyxl.load_workbook(p, read_only=True)
        except Exception:
            continue
        for ws in wb.worksheets:
            for r in range(2, ws.max_row + 1):
                u = ws.cell(r, I_URL + 1).value
                if u and norm_url(u) == target:
                    hits.append((p, ws.title, r, ws.cell(r, I_NAME + 1).value))
        wb.close()
    return hits


def add_cmd(args):
    name = (args.name or '').strip()
    url = norm_url(args.url)
    if not name or not url:
        print('✗ 名称与 URL 必填')
        return 2
    if not os.path.exists(args.table):
        print(f'✗ 目标表不存在: {args.table}')
        return 1

    # 全库查重
    if not args.force:
        hits = find_dup_elsewhere(args.dir, url, args.table)
        if hits:
            print('⚠️ 该 URL 已在其他表收录，已拒绝插入（--force 强制插入）:')
            for p, s, r, n in hits:
                print(f'  {os.path.basename(p)}[{s}] 第{r}行  {n or ""}')
            return 1

    wb = openpyxl.load_workbook(args.table)
    ws = wb[args.sheet] if args.sheet else wb.active
    rows = load_rows(ws)

    # 表内查重
    if not args.force:
        idx = find_row_by_url(rows, url)
        if idx is not None:
            print(f'⚠️ 该 URL 已在当前表[{ws.title}]第{idx+2}行收录（--force 强制重复插入）')
            wb.close()
            return 1

    # --infer 推断缺的类别/类型
    cat, typ = args.cat, args.type
    desc = args.desc
    if args.infer:
        try:
            sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
            from infer_from_url import InferEngine
            eng = InferEngine(args.dir)
            t, c, _ = eng.infer(url)
            typ = typ or t
            cat = cat or c
        except Exception as e:
            print(f'  (infer 失败，跳过: {e})')

    now = datetime.now().strftime('%Y-%m-%d')
    new_row = [None] * COLUMN_COUNT
    new_row[I_NAME] = name
    new_row[I_URL] = url
    new_row[I_CAT] = cat
    new_row[I_TYPE] = typ
    new_row[I_DESC] = desc
    new_row[I_DATE] = now

    pos = insert_pos(rows, cat, args.before, args.after)
    rows.insert(pos, new_row)

    bak = backup(args.table, not args.no_backup)
    renumber(rows)
    write_rows(ws, rows)
    wb.save(args.table)
    wb.close()

    print(f'✅ 已新增: {name}（{url}）')
    print(f'  {args.table}[{ws.title}] 第{pos+2}行（类别:{cat or "（未填）"} 日期:{now}）')
    if bak:
        print(f'  备份: {os.path.basename(bak)}')
    print(f'  → 建议跑 style_table.py 恢复样式，再 verify_table.py 验证')
    return 0


def update_cmd(args):
    url = norm_url(args.url)
    if not os.path.exists(args.table):
        print(f'✗ 目标表不存在: {args.table}')
        return 1
    wb = openpyxl.load_workbook(args.table)
    ws = wb[args.sheet] if args.sheet else wb.active
    rows = load_rows(ws)
    idx = find_row_by_url(rows, url)
    if idx is None:
        print(f'✗ 未找到 URL: {url}（{args.table}[{ws.title}]）')
        wb.close()
        return 1

    updates = []
    if args.set:
        for part in args.set.split(','):
            part = part.strip()
            if '=' not in part:
                print(f'✗ 无效的 --set 段（缺 =）: {part}')
                wb.close()
                return 2
            k, v = (x.strip() for x in part.split('=', 1))
            if k not in COLUMNS:
                print(f'✗ 未知列名: {k}（可用列: {"/".join(COLUMNS)}）')
                wb.close()
                return 2
            updates.append((COLUMNS[k], v))
    if not updates:
        print('✗ 至少需要一个 --set 列=值（或直接给 --set 空触发?）')
        wb.close()
        return 2

    for ci, v in updates:
        rows[idx][ci] = v
    bak = backup(args.table, not args.no_backup)
    write_rows(ws, rows)
    wb.save(args.table)
    wb.close()

    print(f'✅ 已修改: {rows[idx][I_NAME]}（{rows[idx][I_URL]}） 第{idx+2}行')
    for ci, v in updates:
        print(f'  {COLUMN_HEADERS_NAME[ci]} = {v}')
    if bak:
        print(f'  备份: {os.path.basename(bak)}')
    print(f'  → 建议跑 style_table.py 恢复样式')
    return 0


COLUMN_HEADERS_NAME = {i: h for h, i in COLUMNS.items()}


def delete_cmd(args):
    url = norm_url(args.url)
    if not os.path.exists(args.table):
        print(f'✗ 目标表不存在: {args.table}')
        return 1
    wb = openpyxl.load_workbook(args.table)
    ws = wb[args.sheet] if args.sheet else wb.active
    rows = load_rows(ws)
    idx = find_row_by_url(rows, url)
    if idx is None:
        print(f'✗ 未找到 URL: {url}（{args.table}[{ws.title}]）')
        wb.close()
        return 1

    row = rows[idx]
    print('将删除:')
    print(f'  {row[I_NAME]}（{row[I_URL]}）')
    for ci in range(1, COLUMN_COUNT):
        if ci != I_URL and ci != I_NAME and row[ci]:
            print(f'  {COLUMN_HEADERS_NAME[ci]}: {row[ci]}')

    if args.dry_run:
        print('--dry-run：未写入文件')
        wb.close()
        return 0
    if not args.yes:
        wb.close()
        try:
            ans = input('确认删除？(y/N) ').strip().lower()
        except EOFError:
            print('取消（无终端交互，可加 --yes 跳过确认）')
            return 1
        if ans != 'y':
            print('已取消')
            return 1
        wb = openpyxl.load_workbook(args.table)
        ws = wb[args.sheet] if args.sheet else wb.active
        rows = load_rows(ws)
        idx = find_row_by_url(rows, url)
        if idx is None:
            print('✗ 行不存在（表已变化）')
            wb.close()
            return 1

    del rows[idx]
    bak = backup(args.table, not args.no_backup)
    renumber(rows)
    write_rows(ws, rows)
    wb.save(args.table)
    wb.close()

    print(f'✅ 已删除并重编号（原第{idx+2}行）')
    if bak:
        print(f'  备份: {os.path.basename(bak)}')
    print(f'  → 建议跑 style_table.py 恢复样式，再 verify_table.py 验证')
    return 0


def _make_test_table(path):
    """构造 selftest 用临时表。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'AI应用'
    for c, h in enumerate(COLUMNS, start=1):
        ws.cell(1, c).value = h
    data = [
        [1, '旧条目A', 'https://old-a.com', 'AI应用', '官网·工具', '测试A', '', '', '2026-01-01'],
        [2, '旧条目B', 'https://old-b.com', '工具', '官网·工具', '测试B', '', '', '2026-01-02'],
    ]
    for i, row in enumerate(data, start=2):
        for c, v in enumerate(row, start=1):
            ws.cell(i, c).value = v
    wb.save(path)
    wb.close()


def _selftest():
    import tempfile
    tmpdir = tempfile.mkdtemp()
    path = os.path.join(tmpdir, '测试书签汇总.xlsx')
    _make_test_table(path)
    ok = True

    # add
    rc = add_cmd(_Args(table=path, name='新条目', url='https://new.com', cat='AI应用',
                       type='官网·工具', desc='测试新增', sheet=None, dir=tmpdir,
                       infer=False, force=False, before=None, after=None,
                       no_backup=True))
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    rows = load_rows(ws)
    ok &= rc == 0
    ok &= len(rows) == 3
    # 新行应在 AI应用 类别末尾（插在旧条目A后、旧条目B前？不——旧条目B是"工具"，新条目在AI应用末尾=第2行）
    new = rows[1]
    ok &= new[I_NAME] == '新条目' and norm_url(new[I_URL]) == 'https://new.com'
    ok &= rows[0][SEQ] == 1 and rows[1][SEQ] == 2 and rows[2][SEQ] == 3
    ok &= new[I_DATE] == datetime.now().strftime('%Y-%m-%d')
    wb.close()

    # update
    rc = update_cmd(_Args(table=path, url='https://new.com', set='类别=工具,备注=改过',
                          sheet=None, no_backup=True))
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    rows = load_rows(ws)
    new = rows[find_row_by_url(rows, 'https://new.com')]
    ok &= rc == 0 and new[I_CAT] == '工具' and new[I_NOTE] == '改过'
    wb.close()

    # delete
    rc = delete_cmd(_Args(table=path, url='https://new.com', dry_run=False, yes=True,
                          sheet=None, no_backup=True))
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    rows = load_rows(ws)
    ok &= rc == 0 and len(rows) == 2
    ok &= [r[SEQ] for r in rows] == [1, 2]
    wb.close()

    shutil.rmtree(tmpdir, ignore_errors=True)
    print(f'selftest: {"通过" if ok else "失败"}')
    return 0 if ok else 1


class _Args:
    """selftest 用最小参数容器。"""
    def __init__(self, **kw):
        self.__dict__.update(kw)


def main(argv):
    ap = argparse.ArgumentParser(description='书签条目增删改统一入口')
    sub = ap.add_subparsers(dest='cmd')

    p_add = sub.add_parser('add', help='新增书签（全库查重+类别定位+自动日期）')
    p_add.add_argument('name')
    p_add.add_argument('url')
    p_add.add_argument('--table', required=True, help='目标表 .xlsx')
    p_add.add_argument('--cat', default=None, help='类别（插到同类别末尾）')
    p_add.add_argument('--type', dest='type', default=None, help='网站类型')
    p_add.add_argument('--desc', default=None, help='功能定位')
    p_add.add_argument('--sheet', default=None, help='目标 sheet（多 sheet 文件用）')
    p_add.add_argument('--dir', default='.', help='书签库目录（全库查重用）')
    p_add.add_argument('--infer', action='store_true', help='用 infer_from_url 推断缺的类别/类型')
    p_add.add_argument('--force', action='store_true', help='同 URL 也强制插入')
    p_add.add_argument('--before', default=None, help='插到指定名称行之前')
    p_add.add_argument('--after', default=None, help='插到指定名称行之后')
    p_add.add_argument('--no-backup', action='store_true', help='写前不备份')

    p_upd = sub.add_parser('update', help='按 URL 修改指定列')
    p_upd.add_argument('--table', required=True)
    p_upd.add_argument('--url', required=True)
    p_upd.add_argument('--set', required=True, help='列=值,列=值...')
    p_upd.add_argument('--sheet', default=None)
    p_upd.add_argument('--no-backup', action='store_true')

    p_del = sub.add_parser('delete', help='按 URL 删除')
    p_del.add_argument('--table', required=True)
    p_del.add_argument('--url', required=True)
    p_del.add_argument('--sheet', default=None)
    p_del.add_argument('--dry-run', action='store_true', help='只预览不删')
    p_del.add_argument('--yes', action='store_true', help='跳过确认')
    p_del.add_argument('--no-backup', action='store_true')

    ap.add_argument('--selftest', action='store_true', help='运行内置自检')
    args = ap.parse_args(argv)
    if args.selftest:
        return _selftest()
    if not args.cmd:
        ap.print_help()
        return 2

    if args.cmd == 'add':
        return add_cmd(args)
    if args.cmd == 'update':
        return update_cmd(args)
    return delete_cmd(args)


if __name__ == '__main__':
    sys.exit(main(sys.argv[1:]))
