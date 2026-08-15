#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""按本地书签表格生成浏览器书签 HTML（Netscape bookmark 格式）。

数据源:
  - 内容：全库 `*书签汇总.xlsx`（TABLE_GLOB）。每张单 sheet 表 = 顶层文件夹；
    多 sheet 合并表 = 一个组文件夹 + 每个 sheet 一个子文件夹。
  - 顺序：默认按文件名排序；也可用索引表（含"表"列）定义顺序，--index 指定。

用法:
  python export_bookmarks.py [--dir 库目录] [--index 索引.xlsx] [--out out.html]

输出只读，不写任何表。浏览器导入后即得与表格一致的书签结构。
"""
import argparse
import glob
import io
import os
import sys
import time
from html import escape
from datetime import datetime

import openpyxl

if getattr(sys.stdout, 'encoding', '').lower() != 'utf-8':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
try:
    from config import TABLE_GLOB, COLUMNS, COLUMN_COUNT, is_index, is_multi_sheet
except ImportError:
    TABLE_GLOB = '*书签汇总.xlsx'
    COLUMNS = {'序号': 0, '名称': 1, 'URL': 2, '类别': 3, '网站类型': 4,
               '功能定位': 5, '是否重复': 6, '备注': 7, '添加日期': 8}
    COLUMN_COUNT = 9

I_NAME = COLUMNS.get('名称', 1)
I_URL = COLUMNS.get('URL', 2)
I_DATE = COLUMNS.get('添加日期', 8)


def date_to_timestamp(date_val):
    """添加日期转 Unix 时间戳；datetime/date 直接用，字符串解析，空值用当前时间。"""
    if date_val is None:
        return int(time.time())
    if isinstance(date_val, datetime):
        return int(date_val.timestamp())
    if isinstance(date_val, (int, float)):
        return int(date_val)
    try:
        return int(datetime.strptime(str(date_val).strip(), '%Y-%m-%d').timestamp())
    except (ValueError, TypeError):
        return int(time.time())


def folder_name_from_file(fname):
    """表名 → 文件夹名：AI图片书签汇总.xlsx → AI图片（去 书签汇总.xlsx 后缀）。"""
    stem = fname.replace('书签汇总.xlsx', '')
    return stem if stem else fname.replace('.xlsx', '')


def load_sheet(ws):
    """读一个 sheet 的书签行，返回 [(名称, URL, 时间戳), ...]，只保留 http/https。"""
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        url = r[I_URL]
        name = r[I_NAME]
        if not url or not name:
            continue
        url = str(url).strip()
        if not (url.startswith('http://') or url.startswith('https://')):
            continue
        rows.append((str(name).strip(), url, date_to_timestamp(r[I_DATE])))
    return rows


def collect_tables(dirpath):
    """扫描目录，返回 [(文件夹名, items)] 或 [(组名, [[子文件夹名, items], ...])]。
    判别：组的子文件夹是 list，叶子 items 是 tuple。保持文件名字典序。"""
    files = [p for p in glob.glob(os.path.join(dirpath, TABLE_GLOB))
             if not os.path.basename(p).startswith('~$') and '.bak.' not in p
             and not is_index(p)]
    files.sort()

    folders = []
    for path in files:
        fname = os.path.basename(path)
        if is_multi_sheet(path):
            wb = openpyxl.load_workbook(path, read_only=True)
            group = folder_name_from_file(fname)
            subfolders = []
            for ws in wb.worksheets:
                items = load_sheet(ws)
                if items:
                    subfolders.append([ws.title, items])
            wb.close()
            if subfolders:
                folders.append((group, subfolders))
        else:
            wb = openpyxl.load_workbook(path, read_only=True)
            items = load_sheet(wb.active)
            wb.close()
            if items:
                folders.append((folder_name_from_file(fname), items))
    return folders


def order_by_index(folders, index_path):
    """用索引表（含"表"列）重排。folders 元素若是 (组名, subfolders) 则按组名匹配，否则按文件夹名。"""
    wb = openpyxl.load_workbook(index_path, read_only=True)
    order = []
    for row in wb.active.iter_rows(min_row=2, values_only=True):
        name = row[0]
        if name and name != '合计':
            order.append(str(name).replace('.xlsx', ''))
    wb.close()
    if not order:
        return folders

    by_key = {f[0]: f for f in folders}  # 组匹配组名，单表匹配文件夹名
    out = []
    for key in order:
        if key in by_key:
            out.append(by_key.pop(key))
    out.extend(by_key.values())  # 未列出的追加在末尾
    return out


def render(folders, now_ts):
    """渲染 Netscape bookmark HTML。folders: [(文件夹名, items)] 或 [(组名, [(子文件夹名, items)])]。"""
    lines = []
    lines.append('<!DOCTYPE NETSCAPE-Bookmark-file-1>')
    lines.append('<!-- This is an automatically generated file. DO NOT EDIT! -->')
    lines.append('<META HTTP-EQUIV="Content-Type" CONTENT="text/html; charset=UTF-8">')
    lines.append('<TITLE>Bookmarks</TITLE>')
    lines.append('<H1>Bookmarks</H1>')
    lines.append('<DL><p>')
    lines.append(f'    <DT><H3 ADD_DATE="{now_ts}" LAST_MODIFIED="{now_ts}" PERSONAL_TOOLBAR_FOLDER="true">书签栏</H3>')
    lines.append('    <DL><p>')

    def leaf(pad, name, items):
        lines.append(f'{pad}<DT><H3 ADD_DATE="{now_ts}" LAST_MODIFIED="{now_ts}">{escape(name)}</H3>')
        lines.append(f'{pad}<DL><p>')
        for n, u, ts in items:
            lines.append(f'{pad}    <DT><A HREF="{escape(u)}" ADD_DATE="{ts}">{escape(n)}</A>')
        lines.append(f'{pad}</DL><p>')

    for f in folders:
        if isinstance(f[1][0], list):  # 组的子文件夹是 list，叶子条目是 tuple
            pad = ' ' * 8
            lines.append(f'{pad}<DT><H3 ADD_DATE="{now_ts}" LAST_MODIFIED="{now_ts}">{escape(f[0])}</H3>')
            lines.append(f'{pad}<DL><p>')
            for sub_name, items in f[1]:
                leaf(' ' * 12, sub_name, items)
            lines.append(f'{pad}</DL><p>')
        else:
            leaf(' ' * 8, f[0], f[1])
    lines.append('    </DL><p>')
    lines.append('</DL><p>')
    lines.append('')
    return '\n'.join(lines)


def main(argv):
    ap = argparse.ArgumentParser(description='按本地书签表格生成浏览器书签 HTML')
    ap.add_argument('--dir', default='.', help='书签库目录')
    ap.add_argument('--index', default=None, help='索引表 .xlsx（可选，定义文件夹顺序）')
    ap.add_argument('--out', default='bookmarks_new.html', help='输出文件路径')
    ap.add_argument('--selftest', action='store_true')
    args = ap.parse_args(argv)

    if args.selftest:
        return _selftest()

    folders = collect_tables(args.dir)
    if args.index and os.path.exists(args.index):
        folders = order_by_index(folders, args.index)

    if not folders:
        print(f'⚠️ {args.dir} 下未找到书签表（{TABLE_GLOB}）')
        return 1

    total = 0
    for f in folders:
        if isinstance(f[1][0], list):  # 多 sheet 组
            n = sum(len(items) for _, items in f[1])
            print(f'  {f[0]}/: {n} 条')
            for sub, items in f[1]:
                print(f'    - {sub}: {len(items)} 条')
            total += n
        else:
            n = len(f[1])
            print(f'  {f[0]}: {n} 条')
            total += n

    html = render(folders, int(time.time()))
    with open(args.out, 'w', encoding='utf-8') as fh:
        fh.write(html)
    print(f'\n[完成] {len(folders)} 个文件夹, {total} 条链接')
    print(f'   输出: {args.out}')
    return 0


def _selftest():
    import tempfile, shutil
    tmpdir = tempfile.mkdtemp()
    try:
        # 造两表：单 sheet 表 + 多 sheet 表
        p1 = os.path.join(tmpdir, '工具书签汇总.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '工具'
        for c, h in enumerate(['序号', '名称', 'URL', '类别', '网站类型', '功能定位', '是否重复', '备注', '添加日期'], start=1):
            ws.cell(1, c).value = h
        ws.cell(2, 1).value = 1
        ws.cell(2, 2).value = '工具A'
        ws.cell(2, 3).value = 'https://tool-a.com'
        ws.cell(2, 8).value = '2026-01-01'
        ws.cell(3, 1).value = 2
        ws.cell(3, 2).value = '非链接'
        ws.cell(3, 3).value = 'ftp://x.com'  # 应被过滤
        wb.save(p1)
        wb.close()

        p2 = os.path.join(tmpdir, 'AI书签汇总.xlsx')
        wb = openpyxl.Workbook()
        for sname in ['大模型', 'API']:
            ws = wb.create_sheet(sname)
            for c, h in enumerate(['序号', '名称', 'URL', '类别', '网站类型', '功能定位', '是否重复', '备注', '添加日期'], start=1):
                ws.cell(1, c).value = h
            ws.cell(2, 1).value = 1
            ws.cell(2, 2).value = sname + '条目'
            ws.cell(2, 3).value = f'https://{sname.lower()}.com'
            ws.cell(2, 8).value = '2026-02-02'
        wb.remove(wb.active)  # 去掉默认 sheet
        wb.save(p2)
        wb.close()

        folders = collect_tables(tmpdir)
        # 应有两文件夹：AI（组）在前？默认按文件名排序 → AI书签汇总.xlsx 在 工具书签汇总.xlsx 前
        assert len(folders) == 2, folders
        group, subfolders = folders[0]
        assert group == 'AI' and isinstance(subfolders, list)
        assert [s[0] for s in subfolders] == ['大模型', 'API']
        assert len(subfolders[0][1]) == 1 and subfolders[0][1][0][1] == 'https://大模型.com'
        # 单 sheet 表
        fname, items = folders[1]
        assert fname == '工具' and len(items) == 1, items  # ftp 被过滤
        assert items[0][0] == '工具A' and items[0][1] == 'https://tool-a.com'

        html = render(folders, 1234567890)
        assert '大模型' in html and '工具A' in html and 'ftp://x.com' not in html
        assert html.count('<H3') == 5  # 书签栏 + AI + 大模型/API + 工具
        print('selftest: 通过')
        return 0
    except AssertionError as e:
        print(f'selftest: 失败 — {e}')
        return 1
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)


if __name__ == '__main__':
    sys.exit(main(sys.argv[1:]))
