#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""书签自动健康检查：检测死链/错误，最小侵入写回备注 + 独立日志。

用法:
  python health_check_auto.py <xlsx> [<xlsx>...]          # 只报告，不写表
  python health_check_auto.py <xlsx> --write --log health.txt   # 写回备注 + 记日志
  python health_check_auto.py --dir . --limit 200         # 全库前 200 条（预算跨文件累计，需联网）

写回策略（最小侵入）:
  - 仅死链(404/410)/客户端错误(4xx)/连接失败(ERR) 的行，在备注**追加**
    ` | 检查:YYYY-MM-DD 状态:X`；健康行备注不动
  - 幂等: 先剥旧尾巴再重测；恢复的行剥除尾巴还原备注
  - 类别/网站类型不改（避免和 DEAD_CATEGORY 语义混淆）；检查记录写独立日志

Windows 定时（任务计划程序，每日 9 点）:
  schtasks /create /tn "书签健康检查" /tr "D:\\Python\\python.exe
      \"C:\\Users\\spsz0\\.claude\\skills\\maintain-bookmark-tables\\scripts\\health_check_auto.py\"
      --dir E:\\WorkClaudeCode\\yuque\\maintain-bookmark-tables --limit 300 --write
      --log health.log" /sc daily /st 09:00 /f
"""
import sys, os, re, glob, argparse
from datetime import datetime

if getattr(sys.stdout, 'encoding', '').lower() != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')
try:
    from config import COLUMNS, COLUMN_COUNT, TABLE_GLOB, is_index
except ImportError:
    COLUMNS = {'序号': 0, '名称': 1, 'URL': 2, '类别': 3, '网站类型': 4,
               '功能定位': 5, '是否重复': 6, '备注': 7, '添加日期': 8}
    COLUMN_COUNT = 9
    TABLE_GLOB = '*书签汇总.xlsx'
    is_index = lambda x: False

I_NAME = COLUMNS.get('名称', 1)
I_URL = COLUMNS.get('URL', 2)
I_NOTE = COLUMNS.get('备注', 7)

# 检查尾巴语法: ` | 检查:2026-08-16 状态:404`
CHECK_TAIL_RE = re.compile(r' \| 检查:\d{4}-\d{2}-\d{2} 状态:\S+\s*$')


def strip_check_tail(note):
    """剥掉备注末尾的检查尾巴，返回 (干净备注, 是否剥过)。"""
    if not note:
        return '', False
    s = str(note)
    m = CHECK_TAIL_RE.search(s)
    if m:
        return s[:m.start()].rstrip(), True
    return s, False


def load_entries_multi(path):
    """读一个表的所有 sheet。返回 [(sheet名, 行号, 名称, url)]。"""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True)
    out = []
    for ws in wb.worksheets:
        for r in range(2, ws.max_row + 1):
            name = ws.cell(r, I_NAME + 1).value
            if name is None or str(name).strip() == '':
                continue
            url = str(ws.cell(r, I_URL + 1).value or '').strip()
            if url:
                out.append((ws.title, r, name, url))
    wb.close()
    return out


def collect_files(dirpath):
    return sorted(p for p in glob.glob(os.path.join(dirpath, TABLE_GLOB))
                  if not os.path.basename(p).startswith('~$') and '.bak.' not in p
                  and not is_index(p))


def count_entries_multi(path):
    """轻量统计一个表的书签条数（不检查）。供全局预算分配用。"""
    return len(load_entries_multi(path))


def process_table(path, *, limit=0, skip_set=(), check_one=None, out=None, write=False, workers=8):
    """检查一个表，返回 {change: [(sheet,r,name,url,status,note)], dead, err, skip, total}。"""
    import openpyxl
    from collections import defaultdict
    from concurrent.futures import ThreadPoolExecutor, as_completed

    if out is None:
        out = print
    entries = load_entries_multi(path)
    if limit:
        entries = entries[:limit]

    statuses = {}  # (sheet, r) -> status
    with ThreadPoolExecutor(max_workers=workers) as ex:
        futs = {ex.submit(check_one, url, set(skip_set)): (sheet, r)
                for sheet, r, name, url in entries}
        for fut in as_completed(futs):
            key = futs[fut]
            try:
                statuses[key] = fut.result()[1]
            except Exception:
                statuses[key] = 'ERR:check_one'  # 单条异常不影响整批

    changes = []
    dead = err = skip = 0
    today = datetime.now().strftime('%Y-%m-%d')
    wb = None

    def _want_change(st):
        """死链/客户端错误/连接失败 才写备注尾巴。健康/5xx/跳过 不动。"""
        if isinstance(st, int):
            return st in (404, 410) or (400 <= st < 500)
        return isinstance(st, str) and st.startswith('ERR')

    for (sheet, r), st in sorted(statuses.items()):
        if isinstance(st, int):
            if st in (404, 410) or (400 <= st < 500):
                dead += 1
            elif st < 500:
                skip += 1  # 5xx 不写
        elif st == 'SKIP':
            skip += 1
        else:
            err += 1

        if _want_change(st):
            changes.append((sheet, r, st))

    if not write:
        return {'change': 0, 'dead': dead, 'err': err, 'skip': skip, 'total': len(entries)}

    # 写回：死链/错误加尾巴；健康行剥旧尾巴（状态转好恢复原备注）
    changed_set = {(c[0], c[1]) for c in changes}
    real_changes = 0
    wb = openpyxl.load_workbook(path)
    for sheet, r, st in changes:
        ws = wb[sheet]
        cur = ws.cell(r, I_NOTE + 1).value
        base, _ = strip_check_tail(cur)
        new_note = base + f' | 检查:{today} 状态:{st}'
        if new_note != (str(cur) if cur is not None else ''):
            ws.cell(r, I_NOTE + 1).value = new_note
            real_changes += 1
    for sheet, r, name, url in entries:
        if (sheet, r) in changed_set or _want_change(statuses[(sheet, r)]):
            continue  # 待写行不重复处理；5xx 等不写的也不动
        ws = wb[sheet]
        cur = ws.cell(r, I_NOTE + 1).value
        base, had = strip_check_tail(cur)
        if had:
            ws.cell(r, I_NOTE + 1).value = base
            real_changes += 1
    if real_changes:
        import shutil
        shutil.copy2(path, path + '.bak.xlsx')  # 写前备份，可回滚
        wb.save(path)
        out(f'✅ {os.path.basename(path)}: 写回 {real_changes} 行备注（{today}）')
    else:
        out(f'ℹ️ {os.path.basename(path)}: 无死链/错误，无需写回')
    wb.close()

    return {'change': real_changes, 'dead': dead, 'err': err, 'skip': skip, 'total': len(entries)}


def main(argv):
    ap = argparse.ArgumentParser(description='书签自动健康检查（最小侵入写回备注 + 日志）')
    ap.add_argument('xlsx', nargs='*', help='要检查的表文件')
    ap.add_argument('--dir', default='.', help='书签库目录（配 --limit 用）')
    ap.add_argument('--limit', type=int, default=0, help='限制检查条数（--dir 全库时建议必填）')
    ap.add_argument('--skip', default='', help='逗号分隔跳过的域名')
    ap.add_argument('--workers', type=int, default=8, help='并发检查线程数（默认 8，同 check_urls）')
    ap.add_argument('--write', action='store_true', help='写回备注（默认只报告不写表）')
    ap.add_argument('--log', default=None, help='检查记录追加到此日志文件')
    ap.add_argument('--selftest', action='store_true')
    args = ap.parse_args(argv)

    if args.selftest:
        return _selftest()

    from check_urls import check_one  # 复用真实检查，不重实现

    skip_set = {d.strip().lower() for d in args.skip.split(',') if d.strip()}

    files = list(args.xlsx)
    if not files:
        files = collect_files(args.dir)
        if not files:
            print(f'⚠️ {args.dir} 下未找到书签表')
            return 1
        if not args.limit:
            print('⚠️ 全库检查需指定 --limit。例如 --dir . --limit 200')
            return 2

    if args.log:
        log_dir = os.path.dirname(os.path.abspath(args.log)) or '.'
        os.makedirs(log_dir, exist_ok=True)

    summary = {'dead': 0, 'err': 0, 'change': 0, 'total': 0}
    # --dir 全库扫描：--limit 是**全库预算**（跨文件累计，同 check_urls 语义），
    # 先统计各文件条数分配剩余预算，预算耗尽即停，避免某次运行请求过多。
    remaining = args.limit if not args.xlsx else 0
    for f in files:
        if not os.path.exists(f):
            print(f'⚠️ 文件不存在: {f}')
            continue
        if remaining:
            n = count_entries_multi(f)
            if n == 0:
                continue
            take = min(n, remaining)
            res = process_table(f, limit=take, skip_set=skip_set,
                                check_one=check_one, write=args.write, workers=args.workers,
                                out=None if not args.log else
                                (lambda s: _log_and_print(args.log, s)))
            remaining -= take
        else:
            res = process_table(f, skip_set=skip_set, check_one=check_one,
                                write=args.write, workers=args.workers,
                                out=None if not args.log else
                                (lambda s: _log_and_print(args.log, s)))
        for k in summary:
            summary[k] += res.get(k, 0)
        if remaining <= 0:
            break

    print(f"\n汇总: 检查 {summary['total']} 条，死链 {summary['dead']}，连接失败 {summary['err']}，写回 {summary['change']}")
    if args.log:
        _log_and_print(args.log, f"汇总: 检查 {summary['total']} 条，死链 {summary['dead']}，连接失败 {summary['err']}，写回 {summary['change']}")
        print(f'  日志: {args.log}')
    return 1 if (summary['dead'] or summary['err']) else 0


def _log_and_print(log_path, s):
    print(s)
    with open(log_path, 'a', encoding='utf-8') as fh:
        fh.write(f'[{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}] {s}\n')


def _selftest():
    ok = True
    # strip_check_tail 幂等
    s, had = strip_check_tail('备注文字 | 检查:2026-08-16 状态:404')
    ok &= had is True and s == '备注文字'
    s2, had2 = strip_check_tail('备注文字')
    ok &= had2 is False and s2 == '备注文字'
    ok &= strip_check_tail('备注 | 检查:2026-08-16 状态:404 | 再追加')[1] is False  # 只有末尾精确匹配

    # process_table：mock check_one 全死链 + 恢复
    import tempfile, shutil, openpyxl
    tmpdir = tempfile.mkdtemp()
    try:
        p = os.path.join(tmpdir, '测试书签汇总.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '测试'
        for c, h in enumerate(['序号', '名称', 'URL', '类别', '网站类型', '功能定位', '是否重复', '备注', '添加日期'], start=1):
            ws.cell(1, c).value = h
        ws.cell(2, 1).value = 1; ws.cell(2, 2).value = '死链'; ws.cell(2, 3).value = 'https://dead.com'
        ws.cell(2, 8).value = '已有备注'
        ws.cell(3, 1).value = 2; ws.cell(3, 2).value = '健康'; ws.cell(3, 3).value = 'https://ok.com'
        ws.cell(3, 8).value = '正常备注'
        wb.save(p); wb.close()

        fake = lambda url, skip: (url, 404) if 'dead' in url else (url, 200)
        res = process_table(p, check_one=fake, write=False)
        # 只报告不写表
        wb = openpyxl.load_workbook(p)
        assert wb.active.cell(2, 8).value == '已有备注'  # 未写
        wb.close()
        assert res['dead'] == 1 and res['total'] == 2 and res['change'] == 0

        # 写回 + 幂等：跑两遍备注不重复；写前有备份
        res = process_table(p, check_one=fake, write=True)
        assert res['change'] == 1
        assert os.path.exists(p + '.bak.xlsx')  # 写前备份
        res2 = process_table(p, check_one=fake, write=True)
        assert res2['change'] == 0  # 尾巴已存在，strip 后重加 → 无变化
        wb = openpyxl.load_workbook(p)
        note = wb.active.cell(2, 8).value
        assert '检查:' in note and note.count('状态:404') == 1, note
        # 健康行未被加尾巴
        assert '检查:' not in str(wb.active.cell(3, 8).value)
        wb.close()

        # 恢复：死链变健康 → 剥尾巴还原备注
        fake_ok = lambda url, skip: (url, 200)
        res3 = process_table(p, check_one=fake_ok, write=True)
        wb = openpyxl.load_workbook(p)
        assert wb.active.cell(2, 8).value == '已有备注', wb.active.cell(2, 8).value
        wb.close()
        print('selftest: 通过')
        return 0
    except AssertionError as e:
        print(f'selftest: 失败 — {e}')
        return 1
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)


if __name__ == '__main__':
    sys.exit(main(sys.argv[1:]))
