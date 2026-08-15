#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""书签 URL 自动推断：给 URL / 表批量生成 网站类型 + 类别 建议。

用法:
  python infer_from_url.py <url> [...]                  # 给单个/多个 URL 出建议
  python infer_from_url.py <xlsx> --enrich [--changed]  # 给表里所有行出建议清单（不改表）
  python infer_from_url.py --selftest                   # 内置自检（不依赖库目录）

推断优先级:
  1. 内置知名平台规则（网站类型值全部取自库中已有词汇，不造新词）
  2. 从现有库学习: 域名 -> 网站类型/类别 多数派（--dir 指向书签目录，默认当前目录）
  3. 兜底: 官网·工具（库中最高频）+ 类别建议留空

说明: 推断是"建议"，最终类别/类型由人工或 LLM 确认；改表请走正常新增/编辑流程。
"""
import sys, io, os, re, glob
import argparse
from collections import Counter, defaultdict
from urllib.parse import urlparse

# 注意: stdout 包装放在 main() 里执行，被其他脚本 import（如 import_html）时不重复包装

try:
    from config import BUILTIN_RULES, FALLBACK_TYPE, TABLE_GLOB, COLUMN_COUNT, COLUMNS
except ImportError:  # 单文件运行时内联默认
    COLUMN_COUNT = 9
    COLUMNS = {'序号': 0, '名称': 1, 'URL': 2, '类别': 3, '网站类型': 4,
               '功能定位': 5, '是否重复': 6, '备注': 7, '添加日期': 8}
    FALLBACK_TYPE = '官网·工具'
    TABLE_GLOB = '*书签汇总.xlsx'
    # 内置知名平台规则（config.py 中完整版，此处为兜底）
    BUILTIN_RULES = [
        ('gist.github.com', 'GitHub·Gist', '代码分享'),
        ('.github.io', 'GitHub·个人主页', '个人站点'),
        ('github.com', 'GitHub·工具仓库', '开源工具'),
        ('.feishu.cn', '飞书·云文档', '学习教程'),
        ('x.com', 'X·推文', '社交媒体'),
        ('twitter.com', 'X·推文', '社交媒体'),
        ('linux.do', '社区·论坛', '论坛社区'),
        ('bilibili.com', '教程·B站视频', '教程文档'),
        ('youtube.com', '油管·视频教程', '学习教程'),
        ('zhihu.com', '知乎·专栏', '学习教程'),
        ('csdn.net', 'CSDN·博客', '教程博客'),
        ('localhost', '本地·工具', '本地开发'),
        ('127.0.0.1', '本地·工具', '本地开发'),
        ('openai.com', '官网·平台', 'AI聊天'),
        ('chatgpt.com', '官网·平台', 'AI聊天'),
        ('anthropic.com', '官网·平台', 'AI聊天'),
        ('google.com', '官网·搜索', '搜索工具'),
        ('vercel.app', '部署平台·静态托管', '部署平台'),
        ('netlify.app', '部署平台·静态托管', '部署平台'),
    ]

I_NAME = COLUMNS.get('名称', 1)
I_URL = COLUMNS.get('URL', 2)
I_CAT = COLUMNS.get('类别', 3)
I_TYPE = COLUMNS.get('网站类型', 4)


def normalize_host(url):
    """URL -> 小写去 www 的 host；解析失败返回 ''。"""
    try:
        host = (urlparse(url.strip()).hostname or '').lower()
        if host.startswith('www.'):
            host = host[4:]
        return host
    except Exception:
        return ''


class InferEngine:
    """内置规则 + 库学习 的推断器。"""

    def __init__(self, directory='.', online=False):
        # 内置规则: 精确域名 / 后缀 -> (类型, 类别)，后缀按长度降序
        self.exact = {}     # host -> (type, cat)
        self.suffix = []    # [(suffix, type, cat)] 后缀以 '.' 开头
        for suffix, typ, cat in BUILTIN_RULES:
            if suffix.startswith('.'):
                self.suffix.append((suffix, typ, cat))
            else:
                self.exact[suffix] = (typ, cat)
        self.suffix.sort(key=lambda x: -len(x[0]))
        # 库学习: domain -> (类型Counter, 类别Counter)
        self.learn = defaultdict(lambda: (Counter(), Counter()))
        # 表学习: domain -> Counter(表完整文件名)，用于 suggest_table（新增书签进哪张表）
        self.table_learn = defaultdict(Counter)
        # online: 兜底时联网抓 <title> 作为证据（可选，默认离线快）
        self.online = online
        self.load_library(directory)

    def load_library(self, directory):
        """扫描目录所有 *书签汇总.xlsx，学习 域名 -> 类型/类别 多数派。"""
        try:
            import openpyxl
        except ImportError:
            return
        files = sorted(p for p in glob.glob(os.path.join(directory, TABLE_GLOB))
                       if not os.path.basename(p).startswith('~$'))
        for f in files:
            label = os.path.basename(f)  # 表完整文件名，如 AI书签汇总.xlsx（多 sheet 则按 sheet 归并）
            try:
                wb = openpyxl.load_workbook(f, read_only=True)
                for ws in wb.worksheets:  # 多 sheet 文件（如 AI书签汇总.xlsx）逐 sheet 学习
                    for r in range(2, ws.max_row + 1):
                        url = ws.cell(r, I_URL + 1).value
                        typ = ws.cell(r, I_TYPE + 1).value
                        cat = ws.cell(r, I_CAT + 1).value
                        host = normalize_host(str(url)) if url else ''
                        if not host:
                            continue
                        if typ:
                            self.learn[host][0][str(typ).strip()] += 1
                        if cat:
                            self.learn[host][1][str(cat).strip()] += 1
                        self.table_learn[host][label] += 1
                wb.close()
            except Exception:
                continue
        self.learn_loaded = bool(files)

    def _learned(self, host):
        """返回 (类型, 类别) 或 None —— 学习多数派，支持子域名最长后缀。"""
        if not self.learn:
            return None
        keys = [host] + [s for s in self.learn if s.startswith('.') and host.endswith(s)]
        keys = sorted(keys, key=len, reverse=True)
        for k in keys:
            if k not in self.learn:
                continue
            tc, cc = self.learn[k]
            if tc:
                typ = tc.most_common(1)[0][0]
                cat = cc.most_common(1)[0][0] if cc else ''
                return (typ, cat)
        return None

    def infer(self, url):
        """返回 (网站类型建议, 类别建议, 来源)。来源 ∈ 内置/库学习×N/兜底。"""
        host = normalize_host(url)
        if not host:
            return (FALLBACK_TYPE, '', '无法解析')
        # 1. 内置规则：先精确，再祖先域名（chat.openai.com -> openai.com），最后后缀
        if host in self.exact:
            typ, cat = self.exact[host]
            return (typ, cat, '内置')
        parts = host.split('.')
        for i in range(1, len(parts)):
            anc = '.'.join(parts[i:])
            if anc in self.exact:
                typ, cat = self.exact[anc]
                return (typ, cat, '内置')
        for suffix, typ, cat in self.suffix:
            if host.endswith(suffix):
                return (typ, cat, '内置')
        # 2. 库学习
        learned = self._learned(host)
        if learned:
            typ, cat = learned
            return (typ, cat, '库学习')
        # 3. 兜底（--online 时联网抓 <title> 作证据，不强行下结论）
        if self.online:
            title = self._fetch_title(url)
            if title:
                return (FALLBACK_TYPE, '', f'兜底·title:{title[:40]}')
        return (FALLBACK_TYPE, '', '兜底')

    def suggest_table(self, url):
        """返回 (建议表完整文件名, 依据条数) 或 (None, 0)。基于库中 域名->表 多数派。"""
        host = normalize_host(url)
        if not host or not self.table_learn:
            return (None, 0)
        keys = [host] + [s for s in self.table_learn
                         if s.startswith('.') and host.endswith(s)]
        keys = sorted(keys, key=len, reverse=True)
        for k in keys:
            if k not in self.table_learn:
                continue
            cnt = self.table_learn[k]
            if cnt:
                table, n = cnt.most_common(1)[0]
                return (table, n)
        return (None, 0)

    @staticmethod
    def _fetch_title(url):
        """GET 抓 <title> 或 meta og:title 作为兜底证据；失败返回 ''。"""
        import urllib.request
        import re
        try:
            req = urllib.request.Request(
                url, headers={'User-Agent': 'Mozilla/5.0 infer_from_url'},
                method='GET')
            with urllib.request.urlopen(req, timeout=6) as resp:
                raw = resp.read(8192)
                ct = resp.headers.get('Content-Type', '')
                m = re.search(r'charset=([\w-]+)', ct)
                charset = m.group(1) if m else None
        except Exception:
            return ''
        for enc in ([charset] if charset else []) + ['utf-8', 'gbk']:
            try:
                html = raw.decode(enc)
                break
            except (UnicodeDecodeError, LookupError):
                continue
        else:
            html = raw.decode('utf-8', errors='replace')
        m = re.search(r'<title[^>]*>(.*?)</title>', html, re.I | re.S)
        if m:
            return re.sub(r'\s+', ' ', m.group(1)).strip()
        m = re.search(
            r'<meta[^>]+property=["\']og:title["\'][^>]+content=["\']([^"\']*)["\']',
            html, re.I)
        if not m:
            m = re.search(
                r'<meta[^>]+content=["\']([^"\']*)["\'][^>]+property=["\']og:title["\']',
                html, re.I)
        if m:
            return m.group(1).strip()
        return ''


def main(argv):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    ap = argparse.ArgumentParser(description='书签 URL 自动推断 网站类型/类别')
    ap.add_argument('inputs', nargs='*', help='URL 或 xlsx 表路径')
    ap.add_argument('--enrich', action='store_true', help='表模式: 给表每行生成建议清单（不改表）')
    ap.add_argument('--changed', action='store_true', help='表模式: 只显示建议与当前不同的行')
    ap.add_argument('--table', action='store_true', help='URL 模式: 额外给出"应进哪张表"建议')
    ap.add_argument('--online', action='store_true', help='兜底时联网抓 <title> 作证据（默认离线）')
    ap.add_argument('--dir', default='.', help='书签库目录（学习用），默认当前目录')
    ap.add_argument('--selftest', action='store_true', help='内置自检')
    args = ap.parse_args(argv)

    engine = InferEngine(args.dir, online=args.online)

    if args.selftest:
        cases = [
            ('https://github.com/foo/bar', 'GitHub·工具仓库'),
            ('https://skills.sh/xxx', '官网·技能市场'),
            ('https://linux.do/t/123', '社区·论坛'),
            ('https://a.feishu.cn/doc/1', '飞书·云文档'),
            ('https://www.x.com/user', 'X·推文'),
            ('https://blog.csdn.net/foo', 'CSDN·博客'),
            ('http://127.0.0.1:8080', '本地·工具'),
            ('https://chat.openai.com/', '官网·平台'),
            ('https://gist.github.com/u/1', 'GitHub·Gist'),
        ]
        failed = 0
        for url, expect in cases:
            typ, cat, src = engine.infer(url)
            ok = typ == expect
            failed += 0 if ok else 1
            print(f"{'✅' if ok else '❌'} {url} -> {typ}（{src}） 期望 {expect}")
        print(f'selftest: {"通过" if failed == 0 else f"{failed} 处失败"}')
        return 0 if failed == 0 else 1

    if not args.inputs:
        print('用法: python infer_from_url.py <url> [<url>...] 或 <xlsx> --enrich')
        return 2

    # 表模式
    if args.enrich:
        import openpyxl
        for path in args.inputs:
            wb = openpyxl.load_workbook(path, read_only=True)
            ws = wb.active
            print(f'\n== {os.path.basename(path)} ==')
            print(f'{"序号":<4}{"名称":<40}{"URL":<45}{"类型":<16}{"类别":<12}来源')
            for r in range(2, ws.max_row + 1):
                name = ws.cell(r, I_NAME + 1).value
                if name is None:
                    continue
                url = str(ws.cell(r, I_URL + 1).value or '').strip()
                cur_t, cur_c = ws.cell(r, I_TYPE + 1).value, ws.cell(r, I_CAT + 1).value
                sug_t, sug_c, src = engine.infer(url)
                if args.changed and sug_t == str(cur_t).strip() and sug_c == str(cur_c).strip():
                    continue
                print(f'{str(ws.cell(r, 1).value):<4}{str(name)[:38]:<40}{url[:43]:<45}{sug_t:<16}{sug_c:<12}{src}')
            wb.close()
        return 0

    # URL 模式
    for url in args.inputs:
        typ, cat, src = engine.infer(url)
        print(f'{url}\n  网站类型: {typ}\n  类别: {cat or "（待定）"}\n  来源: {src}')
        if args.table:
            table, n = engine.suggest_table(url)
            ttxt = table if table else '（库中未见此域名）'
            ttxt += f'（{n} 条依据）' if n else ''
            print(f'  建议表: {ttxt}')
        print()
    return 0


if __name__ == '__main__':
    sys.exit(main(sys.argv[1:]))
