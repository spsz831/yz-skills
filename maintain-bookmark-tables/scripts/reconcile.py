#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""浏览器书签 ↔ 书签表 双向同步。

方向:
  browser → table: 浏览器导出的 bookmarks.html 里有的、表格里没有的 → 补进表
                    （reconcile.py <html> --dir . [--sync]）
  table    → browser: 表格 → 浏览器书签 HTML（复用 export_bookmarks）
                    （reconcile.py --export --dir . --out bookmarks_sync.html）

用法:
  python reconcile.py bookmarks.html --dir .            # 只出 diff，不写表
  python reconcile.py bookmarks.html --dir . --sync     # 浏览器→表（only_in_html 补入）
  python reconcile.py --export --out bookmarks_sync.html # 表→浏览器

--sync 决策链（补入哪张表）:
  1. 文件夹路径 → 库中表名模糊匹配（AI/ 下的进 AI书签汇总.xlsx）
  2. 兜底 InferEngine.suggest_table(url)
  3. 仍无 → 默认表 DEFAULT_TABLE

说明: --sync 写表前自动备份；只补 only_in_html，不改已在表里的行。
"""
import sys, io, os, glob, argparse

if getattr(sys.stdout, 'encoding', '').lower() != 'utf-8':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
try:
    from config import TABLE_GLOB, is_index, FALLBACK_TYPE
except ImportError:
    TABLE_GLOB = '*书签汇总.xlsx'
    is_index = lambda x: False
    FALLBACK_TYPE = '官网·工具'

DEFAULT_TABLE = '其他书签汇总.xlsx'


def parse_html_entries(path):
    """解析浏览器书签 HTML。返回 [(folder_path_list, name, url)]。复用 import_html.BookmarkParser。"""
    from import_html import BookmarkParser
    with open(path, encoding='utf-8', errors='replace') as fh:
        html = fh.read()
    p = BookmarkParser()
    p.feed(html)
    return [(path_list, name, url) for path_list, name, url in p.bookmarks]


def collect_urls_from_tables(dirpath):
    """读全库所有 URL（含多 sheet），返回 set 去重。"""
    import openpyxl
    from config import COLUMNS
    i_url = COLUMNS.get('URL', 2)
    urls = set()
    for f in glob.glob(os.path.join(dirpath, TABLE_GLOB)):
        if os.path.basename(f).startswith('~$') or '.bak.' in f or is_index(f):
            continue
        wb = openpyxl.load_workbook(f, read_only=True)
        for ws in wb.worksheets:
            for r in ws.iter_rows(min_row=2, values_only=True):
                u = r[i_url]
                if u:
                    urls.add(str(u).strip())
        wb.close()
    return urls


def folder_to_table(path_list, dirpath, default=DEFAULT_TABLE):
    """文件夹路径 → 目标表。文件夹名与库中表名（去 书签汇总.xlsx）精确匹配，最深的优先。"""
    tables = [os.path.basename(p) for p in glob.glob(os.path.join(dirpath, TABLE_GLOB))
              if not is_index(p) and '.bak.' not in os.path.basename(p)]
    for folder in reversed(path_list):  # 最深文件夹优先
        for t in tables:
            if t.replace('书签汇总.xlsx', '') == folder:
                return t
    return default


def compute_diff(html_entries, table_urls):
    """返回 {'only_html': [...], 'only_table': [...], 'common': N}。"""
    only_html = []
    seen = set()
    for path_list, name, url in html_entries:
        u = url.strip()
        if not u:
            continue
        if u in table_urls or u in seen:
            continue
        seen.add(u)
        only_html.append((path_list, name, u))
    common = len(table_urls & seen) + (len(html_entries) - len(only_html))
    only_table = table_urls - seen
    return {'only_html': only_html, 'only_table': only_table, 'common': common}


def apply_html_to_tables(only_html, dirpath, default=DEFAULT_TABLE, dry_run=True):
    """only_in_html 补入表。复用 entry.add_cmd（自动查重/备份/重编号）。"""
    from argparse import Namespace
    from entry import add_cmd

    added = skipped = 0
    for path_list, name, url in only_html:
        table = os.path.join(dirpath, folder_to_table(path_list, dirpath, default))
        if not os.path.exists(table):
            print(f'  ⚠️ 目标表不存在，跳过: {table}（{name}）')
            skipped += 1
            continue
        if dry_run:
            print(f'  [diff] → {os.path.basename(table)}: {name}（{url}）')
            added += 1
            continue
        ns = Namespace(name=name, url=url, table=table,
                       cat=None, type=None, desc=None, sheet=None,
                       infer=True, force=False, before=None, after=None,
                       no_backup=False, dir=dirpath)
        rc = add_cmd(ns)
        if rc == 0:
            added += 1
        else:
            skipped += 1
    return {'added': added, 'skipped': skipped}


def main(argv):
    ap = argparse.ArgumentParser(description='浏览器书签 ↔ 书签表 双向同步')
    ap.add_argument('html', nargs='?', default=None, help='浏览器导出的 bookmarks.html')
    ap.add_argument('--dir', default='.', help='书签库目录')
    ap.add_argument('--default', default=DEFAULT_TABLE, help='兜底目标表')
    ap.add_argument('--sync', action='store_true', help='浏览器→表 补入 only_in_html（默认只出 diff）')
    ap.add_argument('--export', action='store_true', help='表→浏览器 导出 bookmarks_sync.html')
    ap.add_argument('--out', default='bookmarks_sync.html', help='--export 输出路径')
    ap.add_argument('--selftest', action='store_true')
    args = ap.parse_args(argv)

    if args.selftest:
        return _selftest()

    if args.export:
        from export_bookmarks import collect_tables, render
        import time
        folders = collect_tables(args.dir)
        if not folders:
            print(f'⚠️ {args.dir} 下未找到书签表')
            return 1
        html = render(folders, int(time.time()))
        with open(args.out, 'w', encoding='utf-8') as fh:
            fh.write(html)
        print(f'✅ 已导出 {args.out}')
        return 0

    if not args.html:
        ap.error('需指定 bookmarks.html，或用 --export 走表→浏览器方向')
        return 2

    html_entries = parse_html_entries(args.html)
    table_urls = collect_urls_from_tables(args.dir)
    diff = compute_diff(html_entries, table_urls)

    print(f'浏览器书签: {len(html_entries)} 条 | 表内 URL: {len(table_urls)} 个')
    print(f'仅浏览器有: {len(diff["only_html"])} | 仅表内有: {len(diff["only_table"])} | 共有: {diff["common"]}')
    for path_list, name, url in diff['only_html'][:30]:
        folder = '/'.join(path_list) or '(根)'
        print(f'  [browser] {folder} | {name} | {url}')
    if len(diff['only_html']) > 30:
        print(f'  … 其余 {len(diff["only_html"]) - 30} 条')
    for u in list(diff['only_table'])[:10]:
        print(f'  [table]   {u}')
    if len(diff['only_table']) > 10:
        print(f'  … 其余 {len(diff["only_table"]) - 10} 个')

    if args.sync and diff['only_html']:
        print(f'\n补入 only_in_html（写前自动备份）…')
        res = apply_html_to_tables(diff['only_html'], args.dir, args.default, dry_run=False)
        print(f'  ✅ 新增 {res["added"]}，跳过 {res["skipped"]}')
    elif args.sync:
        print('\n无需要补入的条目')
    else:
        print('\n（diff 模式，未写表。加 --sync 补入 only_in_html，--export 导出表→浏览器）')
    return 0


def _selftest():
    import tempfile, shutil, openpyxl
    ok = True
    tmpdir = tempfile.mkdtemp()
    try:
        # 造一个表 + 一个浏览器 html
        p = os.path.join(tmpdir, 'AI书签汇总.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'AI工具'
        for c, h in enumerate(['序号', '名称', 'URL', '类别', '网站类型', '功能定位', '是否重复', '备注', '添加日期'], start=1):
            ws.cell(1, c).value = h
        ws.cell(2, 1).value = 1; ws.cell(2, 2).value = '已有'; ws.cell(2, 3).value = 'https://have.com'
        ws.cell(2, 4).value = 'AI工具'; ws.cell(2, 5).value = '官网·工具'
        wb.save(p); wb.close()

        html_path = os.path.join(tmpdir, 'bookmarks.html')
        with open(html_path, 'w', encoding='utf-8') as fh:
            fh.write('''<!DOCTYPE NETSCAPE-Bookmark-file-1>
<TITLE>B</TITLE><H1>B</H1><DL><p>
<DT><H3>AI</H3><DL><p>
<DT><A HREF="https://have.com">已有</A>
<DT><A HREF="https://new1.com">新书签1</A>
<DT><A HREF="https://new2.com">新书签2</A>
</DL><p></DL><p>''')

        entries = parse_html_entries(html_path)
        assert len(entries) == 3, entries
        assert entries[1][0] == ['AI'] and entries[1][2] == 'https://new1.com'

        table_urls = collect_urls_from_tables(tmpdir)
        assert table_urls == {'https://have.com'}, table_urls

        diff = compute_diff(entries, table_urls)
        assert diff['common'] == 1, diff
        assert len(diff['only_html']) == 2, diff
        assert 'https://have.com' in diff['only_table']

        # folder_to_table：AI 文件夹 → AI书签汇总.xlsx
        t = folder_to_table(['AI'], tmpdir, '其他书签汇总.xlsx')
        assert t == 'AI书签汇总.xlsx', t
        t2 = folder_to_table(['不存在的夹'], tmpdir, '其他书签汇总.xlsx')
        assert t2 == '其他书签汇总.xlsx', t2

        # apply：dry_run 不写表，真实 sync 写一条
        res = apply_html_to_tables(diff['only_html'], tmpdir, dry_run=True)
        assert res['added'] == 2
        wb = openpyxl.load_workbook(p)
        assert wb.active.max_row == 2  # 未写
        wb.close()
        res = apply_html_to_tables(diff['only_html'][:1], tmpdir, dry_run=False)
        assert res['added'] == 1
        wb = openpyxl.load_workbook(p)
        rows = [(wb.active.cell(r, 2).value, wb.active.cell(r, 3).value) for r in range(2, wb.active.max_row + 1)]
        assert ('新书签1', 'https://new1.com') in rows, rows
        wb.close()
        print('selftest: 通过')
        return 0
    except AssertionError as e:
        print(f'selftest: 失败 — {e}')
        return 1
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)


if __name__ == '__main__':
    sys.exit(main(sys.argv[1:]))
