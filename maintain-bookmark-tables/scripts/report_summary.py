#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""全库统计报告：扫描所有 XX书签汇总.xlsx，输出库规模 / 每表概况 / 跨表重复 /
疑似失效 / 低频类别与类型（供合并或审视）。只读。

用法:
  python report_summary.py [--dir 书签目录] [--out report.md]

默认打印到 stdout；--out 写入文件（UTF-8）。
"""
import sys, io, os, glob, re
import argparse
from collections import Counter, defaultdict

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

REF_RE = re.compile(r'同表『([^』]+)』')


def _clip(items, counter, limit=25):
    """低频清单太长时截断，保留前 limit 项并提示剩余数量。"""
    shown = ', '.join(f'{k}×{counter[k]}' for k in items[:limit])
    rest = len(items) - limit
    if rest > 0:
        shown += f' …等 {rest} 项'
    return shown


def load_rows(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    rows = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, 10)]
        if vals[1] is None:
            continue
        rows.append(vals)
    wb.close()
    return rows


def main(argv):
    ap = argparse.ArgumentParser(description='全库统计报告')
    ap.add_argument('--dir', default='.', help='书签库目录')
    ap.add_argument('--out', default='', help='输出到文件（否则打印 stdout）')
    args = ap.parse_args(argv)

    files = sorted(p for p in glob.glob(os.path.join(args.dir, '*书签汇总.xlsx'))
                   if not os.path.basename(p).startswith('~$'))
    if not files:
        print(f'目录下未找到 *书签汇总.xlsx: {args.dir}')
        return 1

    cat_counter = Counter()
    type_counter = Counter()
    url_owner = defaultdict(list)   # url -> [(表名, 序号, 名称)]
    dead_links = []                 # (表名, 序号, 名称, 备注)
    table_info = []                 # (表名, 条数, 重复数, 最早, 最近)
    all_rows = {}

    for f in files:
        rows = load_rows(f)
        all_rows[f] = rows
        name = os.path.basename(f)
        dup = sum(1 for r in rows if str(r[6]).strip() == '是')
        dates = [str(r[8]).strip() for r in rows if r[8]]
        dates.sort()
        earliest = dates[0] if dates else '-'
        latest = dates[-1] if dates else '-'
        table_info.append((name, len(rows), dup, earliest, latest))
        for r in rows:
            cat = str(r[3]).strip() if r[3] else ''
            typ = str(r[4]).strip() if r[4] else ''
            url = str(r[2]).strip() if r[2] else ''
            note = str(r[7]) if r[7] else ''
            if cat:
                cat_counter[cat] += 1
            if typ:
                type_counter[typ] += 1
            if url:
                url_owner[url].append((name, r[0], r[1]))
            if cat == '失效链接' or '失效' in note:
                dead_links.append((name, r[0], r[1], note[:40]))

    total = sum(x[1] for x in table_info)
    dup_total = sum(x[2] for x in table_info)

    # 跨表同 URL（≥2 张不同表）
    cross_dup = [(u, vs) for u, vs in url_owner.items()
                 if len({v[0] for v in vs}) > 1]

    # 低频类别/类型（≤2 次），这些常是近义词/产品名/应合并项
    low_cats = sorted((c for c, n in cat_counter.items() if n <= 2),
                      key=lambda c: -cat_counter[c])
    low_types = sorted((t for t, n in type_counter.items() if n <= 2),
                       key=lambda t: -type_counter[t])

    L = []
    L.append(f'# 书签库统计报告')
    L.append(f'')
    L.append(f'- **表数**: {len(files)}    **总书签**: {total}    **标记重复**: {dup_total}')
    L.append(f'')
    L.append('## 各表概况（条数 / 重复 / 添加日期范围）')
    L.append('')
    L.append('| 表 | 条数 | 重复 | 最早添加 | 最近添加 |')
    L.append('|---|---|---|---|---|')
    for name, cnt, dup, ear, lat in sorted(table_info, key=lambda x: -x[1]):
        L.append(f'| {name} | {cnt} | {dup} | {ear} | {lat} |')
    L.append('')

    L.append(f'## 跨表同 URL（{len(cross_dup)} 组，核对是否应标重复）')
    L.append('')
    if cross_dup:
        for url, vs in sorted(cross_dup, key=lambda x: -len(x[1])):
            where = ' / '.join(sorted({v[0] for v in vs}))
            L.append(f'- {url}  出现在: {where}')
    else:
        L.append('（无）')
    L.append('')

    L.append(f'## 疑似失效（类别=失效链接 或 备注含"失效"，{len(dead_links)} 条）')
    L.append('')
    if dead_links:
        for name, seq, title, note in dead_links:
            L.append(f'- {name} 序号{seq} {title}  {note}')
    else:
        L.append('（无）')
    L.append('')

    L.append('## 低频类别（≤2 次，多为近义词/待合并/产品名，供审视）')
    L.append('')
    L.append(_clip(low_cats, cat_counter) or '（无）')
    L.append('')

    L.append('## 低频网站类型（≤2 次，多为近义词/笼统值，供审视）')
    L.append('')
    L.append(_clip(low_types, type_counter) or '（无）')
    L.append('')

    L.append('## 类别分布 Top 20')
    L.append('')
    for c, n in cat_counter.most_common(20):
        L.append(f'- {c}: {n}')
    L.append('')

    text = '\n'.join(L)
    if args.out:
        with open(args.out, 'w', encoding='utf-8') as fh:
            fh.write(text)
        print(f'已写入 {args.out}（{len(files)} 张表 / {total} 条）')
    else:
        print(text)
    return 0


if __name__ == '__main__':
    sys.exit(main(sys.argv[1:]))
