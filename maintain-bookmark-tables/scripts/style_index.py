#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""索引表（书签整理最终清单）统一格式化：表头深蓝白字 + 数据行斑马纹 + 合计行突出。

与 style_table.py 同一风格体系（微软雅黑 / 3867A6 / EEF4FB），但针对 4 列索引表：
  列1 表         左对齐+换行
  列2 条数       居中
  列3 所属类别   左对齐+换行（内容最长）
  列4 说明       左对齐+换行

用法: python style_index.py <xlsx> [<xlsx>...]
"""
import sys, io, openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
try:
    from config import STYLE
except ImportError:  # 单文件运行时内联默认
    STYLE = dict(font_name='微软雅黑', header_size=11, body_size=10.5,
                 header_fill='3867A6', accent_fill='EEF4FB', total_fill='FFF2CC',
                 url_color='2563EB', border_color='999999',
                 header_row_height=26, body_row_height=22)

THIN = Side(style='thin', color=STYLE['border_color'])  # 深灰，WPS 中清晰可见
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FONT = Font(name=STYLE['font_name'], size=STYLE['header_size'], bold=True, color='FFFFFF')
BODY_FONT = Font(name=STYLE['font_name'], size=STYLE['body_size'], color='000000')
TOTAL_FONT = Font(name=STYLE['font_name'], size=STYLE['body_size'], bold=True, color='000000')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)


def style_sheet(ws):
    # 表头
    ws.row_dimensions[1].height = STYLE['header_row_height']
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(1, c)
        cell.font = HEADER_FONT
        cell.fill = PatternFill('solid', start_color=STYLE['header_fill'])
        cell.alignment = CENTER
        cell.border = BORDER
    # 数据行（含合计行）
    accent_fill = PatternFill('solid', start_color=STYLE['accent_fill'])
    white_fill = PatternFill('solid', start_color='FFFFFF')
    total_fill = PatternFill('solid', start_color=STYLE['total_fill'])
    for r in range(2, ws.max_row + 1):
        if not any(ws.cell(r, c).value is not None for c in range(1, ws.max_column + 1)):
            continue
        is_total = str(ws.cell(r, 1).value).strip() == '合计'
        ws.row_dimensions[r].height = STYLE['body_row_height']
        fill = total_fill if is_total else (accent_fill if r % 2 == 0 else white_fill)
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            cell.fill = fill
            cell.border = BORDER
            cell.alignment = CENTER if c == 2 else LEFT
            cell.font = TOTAL_FONT if is_total else BODY_FONT
    # 自动筛选 + 冻结首行
    ws.auto_filter.ref = f'A1:{openpyxl.utils.get_column_letter(ws.max_column)}{ws.max_row}'
    ws.freeze_panes = 'A2'


def main(argv):
    for path in argv:
        wb = openpyxl.load_workbook(path)
        for ws in wb.worksheets:
            style_sheet(ws)
        wb.save(path)
        print(f'已统一格式化索引表: {path}')
    return 0


if __name__ == '__main__':
    sys.exit(main(sys.argv[1:]))
