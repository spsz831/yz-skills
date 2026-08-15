#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""书签汇总表统一格式化：字体 / 行高 / 对齐 / 边框 / 斑马纹隔行 / URL 蓝字。

skill 统一样式规范（全表一致，脚本幂等，可反复跑）：
  表头   微软雅黑 11 加粗 白字 深蓝底(3867A6) 居中 行高 26
  数据行 微软雅黑 10.5 细边框 行高 22，浅蓝(EEF4FB)/白 隔行交替
  对齐   序号/类别/网站类型/是否重复/添加日期 居中；名称/URL/功能定位/备注 左对齐+换行
  URL    蓝字(2563EB)
  列宽   序号5 / 名称38 / URL50 / 类别14 / 网站类型14 / 功能定位32 / 是否重复10 / 备注30 / 添加日期12

用法: python style_table.py <xlsx> [<xlsx>...] [--accent EEF4FB]
"""
import sys, io, argparse, openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
try:
    from config import STYLE, COLUMN_HEADERS, COLUMNS, COLUMN_COUNT
except ImportError:  # 单文件运行时内联默认
    STYLE = dict(font_name='微软雅黑', header_size=11, body_size=10.5,
                 header_fill='3867A6', accent_fill='EEF4FB', total_fill='FFF2CC',
                 url_color='2563EB', border_color='999999',
                 header_row_height=26, body_row_height=22,
                 col_widths=[5, 38, 50, 14, 14, 32, 10, 30, 12],
                 center_cols=('序号', '类别', '网站类型', '是否重复', '添加日期'),
                 left_cols=('名称', 'URL', '功能定位', '备注'))
    COLUMN_HEADERS = ['序号', '名称', 'URL', '类别', '网站类型', '功能定位', '是否重复', '备注', '添加日期']
    COLUMNS = {h: i for i, h in enumerate(COLUMN_HEADERS)}
    COLUMN_COUNT = 9

ACCENT_FILL = STYLE['accent_fill']
THIN = Side(style='thin', color=STYLE['border_color'])  # 深灰，WPS 中清晰可见
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FONT = Font(name=STYLE['font_name'], size=STYLE['header_size'], bold=True, color='FFFFFF')
BODY_FONT = Font(name=STYLE['font_name'], size=STYLE['body_size'], color='000000')
URL_FONT = Font(name=STYLE['font_name'], size=STYLE['body_size'], color=STYLE['url_color'])
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
CENTER_COLS = tuple(COLUMNS[c] + 1 for c in STYLE['center_cols'])   # 序号/类别/网站类型/是否重复/添加日期
LEFT_COLS = tuple(COLUMNS[c] + 1 for c in STYLE['left_cols'])        # 名称/URL/功能定位/备注
COL_WIDTHS = tuple(STYLE['col_widths'])  # 9 列标准列宽


def _last_data_row(ws):
    """找最后一个有任一非空单元格的行号（含表头）。"""
    for r in range(ws.max_row, 0, -1):
        if any(ws.cell(r, c).value is not None for c in range(1, ws.max_column + 1)):
            return r
    return 1


def _apply_filter_freeze(ws):
    """顶部第一行自动筛选 + 冻结首行（幂等）。"""
    last = _last_data_row(ws)
    if last >= 1:
        ws.auto_filter.ref = f'A1:{openpyxl.utils.get_column_letter(ws.max_column)}{last}'
    ws.freeze_panes = 'A2'


def style_sheet(ws, accent):
    url_col = COLUMNS.get('URL', 2) + 1  # 1-based URL 列号
    # 列宽
    for c, w in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w
    # 表头
    ws.row_dimensions[1].height = STYLE['header_row_height']
    for c in range(1, COLUMN_COUNT + 1):
        cell = ws.cell(1, c)
        cell.font = HEADER_FONT
        cell.fill = PatternFill('solid', start_color=STYLE['header_fill'])
        cell.alignment = CENTER
        cell.border = BORDER
    # 数据行（整行全空才跳过；非空残行也加框线）
    accent_fill = PatternFill('solid', start_color=accent)
    white_fill = PatternFill('solid', start_color='FFFFFF')
    for r in range(2, ws.max_row + 1):
        if not any(ws.cell(r, c).value is not None for c in range(1, COLUMN_COUNT + 1)):
            continue
        ws.row_dimensions[r].height = STYLE['body_row_height']
        fill = accent_fill if r % 2 == 0 else white_fill
        for c in range(1, COLUMN_COUNT + 1):
            cell = ws.cell(r, c)
            cell.fill = fill
            cell.border = BORDER
            cell.alignment = CENTER if c in CENTER_COLS else LEFT
            cell.font = URL_FONT if c == url_col else BODY_FONT
    _apply_filter_freeze(ws)


def main(argv):
    ap = argparse.ArgumentParser()
    ap.add_argument('xlsx', nargs='+')
    ap.add_argument('--accent', default=ACCENT_FILL, help='斑马纹浅色（十六进制无#）')
    args = ap.parse_args(argv)
    for path in args.xlsx:
        wb = openpyxl.load_workbook(path)
        for ws in wb.worksheets:
            style_sheet(ws, args.accent)
        wb.save(path)
        print(f'已统一格式化: {path}')
    return 0


if __name__ == '__main__':
    # 自检：COL_WIDTHS 必须与 COLUMN_HEADERS 对齐
    assert len(COL_WIDTHS) == COLUMN_COUNT, f'列宽定义应有{COLUMN_COUNT}项: {COL_WIDTHS}'
    sys.exit(main(sys.argv[1:]))
